import json
import os
import sys
import re
import platform
import shutil
import subprocess
import tempfile
import time
import hashlib
import asyncio
import aiohttp
from datetime import datetime, timezone
from pathlib import Path
from typing import Dict, List, Tuple, Optional, Any

import numpy as np
import requests
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from sentence_transformers import SentenceTransformer
from dotenv import load_dotenv

BASE_DIR = Path(__file__).resolve().parent
load_dotenv(BASE_DIR / ".env", override=False)

#
def _is_ubuntu_linux() -> bool:
    """Nhan dien Ubuntu dung hon platform.platform() (tren Ubuntu thuong khong co chuoi 'ubuntu')."""
    if platform.system().lower() != "linux":
        return False
    try:
        info = platform.freedesktop_os_release()
    except (OSError, AttributeError):
        info = {}
    if info.get("ID", "").lower() == "ubuntu":
        return True
    if "ubuntu" in info.get("ID_LIKE", "").lower():
        return True
    try:
        with open("/etc/os-release", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if line.upper().startswith("ID="):
                    val = line.split("=", 1)[1].strip().strip('"').lower()
                    if val == "ubuntu":
                        return True
                if line.upper().startswith("ID_LIKE="):
                    if "ubuntu" in line.lower():
                        return True
    except OSError:
        pass
    return False


def _is_tesla_t4_name(gpu_name: str) -> bool:
    """Tesla T4 / TU104GL — PCI co the ghi TU104GL [Tesla T4]; CUDA name thuong chua T4 hoac TU104."""
    g = gpu_name.lower()
    if "tesla t4" in g:
        return True
    if "tu104" in g:
        return True
    if re.search(r"\bt4\b", g):
        return True
    return False


# ── GPU detection ─────────────────────────────────────────────────────────────
try:
    import torch
    _REQUIRE_UBUNTU_T4 = os.getenv("REQUIRE_UBUNTU_T4", "0").strip().lower() not in ("0", "false", "no")
    _IS_UBUNTU = _is_ubuntu_linux()
    print(f"[GPU] PyTorch version: {torch.__version__}")
    print(f"[GPU] CUDA available: {torch.cuda.is_available()}")
    if hasattr(torch.version, 'cuda'):
        print(f"[GPU] CUDA version (torch built with): {torch.version.cuda}")
    if _REQUIRE_UBUNTU_T4 and not _IS_UBUNTU:
        raise RuntimeError(
            "Yeu cau he dieu hanh Ubuntu de su dung GPU NVIDIA Tesla T4. "
            "Dat REQUIRE_UBUNTU_T4=0 neu muon bo qua rang buoc nay."
        )
    if torch.cuda.is_available():
        nd = torch.cuda.device_count()
        print(f"[GPU] CUDA devices ({nd}) — VMware/console dung GPU khac; day chi la hang NVIDIA:")
        for idx in range(nd):
            print(f"[GPU]   cuda:{idx} = {torch.cuda.get_device_name(idx)}")
        t4_index = None
        for idx in range(nd):
            gpu_name = torch.cuda.get_device_name(idx)
            if _is_tesla_t4_name(gpu_name):
                t4_index = idx
                break
        if _REQUIRE_UBUNTU_T4 and t4_index is None:
            raise RuntimeError(
                "Khong tim thay GPU NVIDIA Tesla T4. "
                "Dat REQUIRE_UBUNTU_T4=0 neu muon cho phep GPU khac/CPU."
            )
        selected_idx = t4_index if t4_index is not None else 0
        torch.cuda.set_device(selected_idx)
        DEVICE = "cuda"
        GPU_NAME = torch.cuda.get_device_name(selected_idx)
        GPU_MEM_GB = torch.cuda.get_device_properties(selected_idx).total_memory / 1e9
    elif hasattr(torch.backends, "mps") and torch.backends.mps.is_available():
        DEVICE = "mps"
        GPU_NAME = "Apple MPS"
        GPU_MEM_GB = 0.0
    else:
        DEVICE = "cpu"
        GPU_NAME = "CPU"
        GPU_MEM_GB = 0.0
        
        # Enforce GPU if requested
        _REQUIRE_GPU = os.getenv("REQUIRE_GPU", "0").strip().lower() not in ("0", "false", "no")
        if _REQUIRE_GPU:
            raise RuntimeError(
                "REQUIRE_GPU=1 but no NVIDIA CUDA or Apple MPS device found! "
                "Check drivers or set REQUIRE_GPU=0."
            )

        if hasattr(torch.version, 'cuda') and torch.version.cuda is None:
            print("[GPU] WARNING: PyTorch is CPU-only build! "
                  "Reinstall with: pip install torch==2.11.0+cu126 --extra-index-url "
                  "https://download.pytorch.org/whl/cu126")
    HAS_TORCH = True
except ImportError:
    import types
    torch = None          # type: ignore
    HAS_TORCH = False
    DEVICE = "cpu"
    GPU_NAME = "CPU (torch not installed)"
    GPU_MEM_GB = 0.0

print(f"[GPU] Device: {DEVICE}  ({GPU_NAME})"
      + (f"  VRAM={GPU_MEM_GB:.1f}GB" if GPU_MEM_GB else ""))

try:
    from tqdm import tqdm as _tqdm
    HAS_TQDM = True
except ImportError:
    HAS_TQDM = False
    class _tqdm:  # type: ignore
        def __init__(self, total=0, desc="", unit="", **kw):
            self.total=total; self.n=0; self.desc=desc
        def update(self, n=1):
            self.n+=n
            pct = int(100*self.n/self.total) if self.total else 0
            print(f"\r  {self.desc}: {self.n}/{self.total} ({pct}%)", end="", flush=True)
        def close(self): print()
        def __enter__(self): return self
        def __exit__(self,*a): self.close()


# =========================
# CONFIG
# =========================
CLAUDE_API_URL = "https://api.anthropic.com/v1/messages"

CLAUDE_MODEL_TRANSLATE   = os.getenv("CLAUDE_MODEL_TRANSLATE",   "claude-sonnet-4-6")
CLAUDE_MODEL_AI_DESC     = os.getenv("CLAUDE_MODEL_AI_DESC",     "claude-sonnet-4-6")
CLAUDE_MODEL_AI_FALLBACK = os.getenv("CLAUDE_MODEL_AI_FALLBACK", "claude-sonnet-4-6")
CLAUDE_MODEL           = os.getenv("CLAUDE_MODEL",            "claude-sonnet-4-6")

ANTHROPIC_VERSION  = "2023-06-01"
ANTHROPIC_API_KEY  = os.getenv("ANTHROPIC_API_KEY", "")

INPUT_FILE         = sys.argv[1] if len(sys.argv) > 1 else os.getenv("INPUT_FILE", "Test_02_to_V365_converted.xlsx")
OUTPUT_FILE        = os.getenv("OUTPUT_FILE",  "mep20.xlsx")
SHEET_NAME         = os.getenv("SHEET_NAME")  or None
KNOWLEDGE_JSON_FILE = os.getenv("KNOWLEDGE_JSON_FILE", "boq_merged.json")
KNOWLEDGE_LOCAL_FILE = os.getenv("KNOWLEDGE_LOCAL_FILE", "boq_merged_local.json")
BOQ_MERGED_JSON_FILE = os.getenv("BOQ_MERGED_JSON_FILE", "boq_merged.json")
BOQ_MERGED_LOCAL_FILE = os.getenv("BOQ_MERGED_LOCAL_FILE", "boq_merged_local.json")

HEADER_ROW   = 1
START_ROW    = 2

HEADER_ITEM_CODE              = "Mã hạng mục"
HEADER_PHAN_MUC               = "Phân mục"
HEADER_ITEM_NAME              = "Tên hạng mục"
HEADER_SOURCE_DESC            = "Mô tả đặc điểm hạng mục"
HEADER_TRANSLATED_ITEM_NAME   = "Dịch tên hạng mục (Nếu có)"
HEADER_TRANSLATED_DESC        = "Dịch thông tin mô tả (Nếu có)"
HEADER_AI_DESC                = "Mô tả AI"
NHAN_HIEU_HEADERS = ["Thương hiệu", "Nhãn hiệu", "Hãng sản xuất", "Hãng", "NSX", "Brand", "Manufacturer"]

HEADER_BEST_MATCH_CODE  = "Mã tương đồng cao nhất"
HEADER_BEST_MATCH_SCORE = "Điểm tương đồng cao nhất"
HEADER_SIM_PASS         = "Nhãn đạt ngưỡng"

# ── Concurrency ───────────────────────────────────────────────────────────────
# Tự động tăng batch size khi có GPU (embed nhanh hơn nhiều)
_GPU_BATCH_MULTIPLIER = 4 if DEVICE in ("cuda", "mps") else 1
BATCH_SIZE              = int(os.getenv("BATCH_SIZE",
    str(100 * _GPU_BATCH_MULTIPLIER)))   # 100 CPU → 400 GPU
MAX_CONCURRENT_API_CALLS= int(os.getenv("MAX_CONCURRENT_API_CALLS", "5"))
REQUEST_TIMEOUT         = int(os.getenv("REQUEST_TIMEOUT",          "120"))
MAX_RETRIES             = int(os.getenv("MAX_RETRIES",               "8"))
MIN_DELAY_BETWEEN_CALLS = float(os.getenv("MIN_DELAY_BETWEEN_CALLS","0.3"))
MAX_RETRY_WAIT          = float(os.getenv("MAX_RETRY_WAIT",         "120.0"))
SAVE_EVERY_N_BATCHES    = int(os.getenv("SAVE_EVERY_N_BATCHES",     "5"))

# ── Embedding batch size cho GPU (lớn hơn nhiều so với CPU) ──────────────────
# CUDA A100: 4096+ an toàn; RTX 3090: 2048; MPS: 512; CPU: 256
_DEFAULT_EMBED_BATCH = (
    4096 if (DEVICE == "cuda" and GPU_MEM_GB > 20) else
    2048 if (DEVICE == "cuda" and GPU_MEM_GB > 8) else
    1024 if (DEVICE == "cuda") else
    512  if (DEVICE == "mps") else
    256
)
EMBED_BATCH_SIZE = int(os.getenv("EMBED_BATCH_SIZE", str(_DEFAULT_EMBED_BATCH)))

FORCE_REPROCESS_ITEM_NAME_VI = os.getenv("FORCE_REPROCESS_ITEM_NAME_VI","false").lower()=="true"
FORCE_REPROCESS_DESC_VI      = os.getenv("FORCE_REPROCESS_DESC_VI",     "false").lower()=="true"
FORCE_REPROCESS_AI           = os.getenv("FORCE_REPROCESS_AI",          "false").lower()=="true"

EMBED_MODEL_NAME          = os.getenv("EMBED_MODEL_NAME", "paraphrase-multilingual-MiniLM-L12-v2")
SIMILARITY_AUDIT_THRESHOLD= float(os.getenv("SIMILARITY_AUDIT_THRESHOLD","0.88"))

INSTANCE_ID = os.getenv("INSTANCE_ID", "local")

PERSIST_DIR                = Path(os.getenv("PERSIST_DIR", ".persist"))

TRANSLATION_CACHE_FILE     = PERSIST_DIR / f"translation_cache_{INSTANCE_ID}.json"
AI_DESC_CACHE_FILE         = PERSIST_DIR / f"ai_desc_cache_{INSTANCE_ID}.json"
SIM_AUDIT_CACHE_FILE       = PERSIST_DIR / f"sim_audit_cache_{INSTANCE_ID}.json"
LEARNED_KNOWLEDGE_FILE     = PERSIST_DIR / f"learned_knowledge_{INSTANCE_ID}.json"
OUTPUT_CACHE_FILE          = PERSIST_DIR / f"output_cache_{INSTANCE_ID}.json"
QUERY_EMBED_CACHE_FILE     = PERSIST_DIR / f"query_embed_cache_{INSTANCE_ID}.json"
QUERY_EMBED_CACHE_NPZ_FILE = PERSIST_DIR / f"query_embed_cache_{INSTANCE_ID}.npz"
CHECKPOINT_FILE            = PERSIST_DIR / f"checkpoint_{INSTANCE_ID}.json"

KNOWLEDGE_CACHE_META_FILE  = PERSIST_DIR / "knowledge_index_meta.json"
KNOWLEDGE_CACHE_VEC_FILE   = PERSIST_DIR / "knowledge_index_vectors.npz"
LEARNED_CACHE_META_FILE    = PERSIST_DIR / "learned_index_meta.json"
LEARNED_CACHE_VEC_FILE     = PERSIST_DIR / "learned_index_vectors.npz"
PURGE_SIGNATURE_FILE       = PERSIST_DIR / "purge_signature.txt"

AUTO_LEARN_ENABLED         = os.getenv("AUTO_LEARN_ENABLED",        "true").lower()=="true"
AUTO_LEARN_MIN_SCORE       = float(os.getenv("AUTO_LEARN_MIN_SCORE","0.88"))
AUTO_LEARN_CODELESS_ROWS   = os.getenv("AUTO_LEARN_CODELESS_ROWS",  "true").lower()=="true"
AUTO_LEARN_USE_AI_DESC     = os.getenv("AUTO_LEARN_USE_AI_DESC",    "true").lower()=="true"
INCLUDE_LEARNED_IN_CURRENT_RUN = os.getenv("INCLUDE_LEARNED_IN_CURRENT_RUN","false").lower()=="true"
REWRITE_AI_IF_HAS_COMPONENTS   = os.getenv("REWRITE_AI_IF_HAS_COMPONENTS",  "true").lower()=="true"

ITEM_CODE_COL_FALLBACK          = None
ITEM_NAME_COL_FALLBACK          = 2
SOURCE_DESC_COL_FALLBACK        = 3
TRANSLATED_ITEM_NAME_COL_FALLBACK=4
TRANSLATED_DESC_COL_FALLBACK    = 5
OUTPUT_AI_COL_FALLBACK          = 6

ROW_HEIGHT_MIN = 42
ROW_HEIGHT_MAX = 220

SESSION = requests.Session()
if ANTHROPIC_API_KEY:
    SESSION.headers.update({
        "x-api-key":          ANTHROPIC_API_KEY,
        "anthropic-version":  ANTHROPIC_VERSION,
        "content-type":       "application/json",
    })

# ── Runtime state ──────────────────────────────────────────────────────────────
_EMBED_MODEL: Optional[Any] = None
KNOWLEDGE_INDEX: Optional[dict] = None

# ── GPU tensor cache cho similarity matrix ────────────────────────────────────
# Lưu knowledge matrix dưới dạng GPU tensor, rebuild khi index thay đổi
_GPU_SIM_MATRIX: Optional[Any] = None   # torch.Tensor trên DEVICE
_GPU_SIM_MATRIX_SIG: str = ""           # signature để detect khi nào rebuild

TRANSLATION_CACHE:  Dict[Tuple, Tuple[str,str]] = {}
AI_DESC_CACHE:      Dict[Tuple, str]            = {}
SIM_AUDIT_CACHE:    Dict[str, dict]             = {}
QUERY_EMBED_CACHE:  Dict[str, Any]              = {}
LEARNED_KEYS_CACHE: set                         = set()
LEARNED_ITEMS_RAM:  List[dict]                  = []
OUTPUT_CACHE:       Dict[str, dict]             = {}

_INST_TRANSLATION_KEYS: set = set()
_INST_AI_DESC_KEYS:     set = set()
_INST_SIM_AUDIT_KEYS:   set = set()
_INST_OUTPUT_CACHE_KEYS:set = set()
_INST_QUERY_EMBED_KEYS: set = set()

LEARNED_AI_LOOKUP:  Dict[str, str]              = {}

_LEARNED_ITEMS_DIRTY  = False
_TRANSLATION_DIRTY    = False
_AI_CACHE_DIRTY       = False
_QUERY_EMBED_DIRTY    = False
_OUTPUT_CACHE_DIRTY   = False


# =========================
# EMBEDDING  (GPU-accelerated)
# =========================
def get_embed_model() -> SentenceTransformer:
    """Load SentenceTransformer lên DEVICE (cuda/mps/cpu)."""
    global _EMBED_MODEL
    if _EMBED_MODEL is not None:
        return _EMBED_MODEL
    print(f"Loading embedding model: {EMBED_MODEL_NAME} → device={DEVICE} ...")
    _EMBED_MODEL = SentenceTransformer(EMBED_MODEL_NAME, device=DEVICE)
    # Nếu torch khả dụng và CUDA, bật half-precision để tăng throughput
    if HAS_TORCH and DEVICE == "cuda":
        try:
            _EMBED_MODEL = _EMBED_MODEL.half()   # FP16 trên GPU
            print(f"  ✓ Embedding model loaded (FP16 on {GPU_NAME})")
        except Exception:
            print(f"  ✓ Embedding model loaded (FP32 on {GPU_NAME})")
    else:
        print(f"  ✓ Embedding model loaded on {DEVICE}")
    return _EMBED_MODEL


# =========================
# BASIC HELPERS
# =========================
def safe_str(value) -> str:
    if value is None: return ""
    return str(value).strip()

def normalize_spaces(text: str) -> str:
    text = safe_str(text)
    if not text: return ""
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    lines = []
    for line in text.split("\n"):
        line = re.sub(r"\s+", " ", line).strip()
        if line: lines.append(line)
    return "\n".join(lines).strip()

def normalize_inline(text: str) -> str:
    return re.sub(r"\s+", " ", safe_str(text)).strip()

def normalize_basic_compare(text: str) -> str:
    text = normalize_spaces(text).lower()
    return re.sub(r"[^\w]+", "", text)

def capitalize_first(text: str) -> str:
    text = normalize_inline(text)
    if not text: return ""
    return text[0].upper() + text[1:]

def dedupe_preserve_order(items: List[str]) -> List[str]:
    seen, result = set(), []
    for item in items:
        key = normalize_basic_compare(item)
        if not key or key in seen: continue
        seen.add(key); result.append(item)
    return result

def build_component_lookup(raw_list: List[dict]) -> Dict[str, List[dict]]:
    lookup: Dict[str, List[dict]] = {}
    for raw in raw_list:
        components = raw.get("components") or []
        if not components:
            continue
        ma_hieu = normalize_ma_hieu(raw.get("ma_hieu"))
        if ma_hieu and ma_hieu not in lookup:
            lookup[ma_hieu] = components
        ten_cong_viec = normalize_inline(raw.get("ten_cong_viec"))
        if ten_cong_viec:
            lookup.setdefault(f"name::{ten_cong_viec.lower()}", components)
    return lookup

def load_boq_component_lookup() -> Dict[str, List[dict]]:
    all_items = []
    for f_path in [BOQ_MERGED_JSON_FILE, BOQ_MERGED_LOCAL_FILE]:
        if os.path.exists(f_path):
            with open(f_path, "r", encoding="utf-8") as f:
                all_items.extend(json.load(f))
    return build_component_lookup(all_items)

def resolve_components_for_item(raw: dict, component_lookup: Dict[str, List[dict]]) -> List[dict]:
    ma_hieu = normalize_ma_hieu(raw.get("ma_hieu"))
    if ma_hieu and ma_hieu in component_lookup:
        return component_lookup[ma_hieu]
    ten_cong_viec = normalize_inline(raw.get("ten_cong_viec"))
    if ten_cong_viec:
        return component_lookup.get(f"name::{ten_cong_viec.lower()}", raw.get("components") or [])
    return raw.get("components") or []

def refresh_row_components_from_boq(row: dict, component_lookup: Dict[str, List[dict]]) -> List[dict]:
    primary_match = row.get("primary_match") or {}
    refreshed = resolve_components_for_item(primary_match, component_lookup)
    if refreshed:
        primary_match["components"] = refreshed
        row["primary_match"] = primary_match
    return refreshed

def format_component_entry(component: Any) -> str:
    if isinstance(component, dict):
        parts = [
            normalize_inline(component.get("ma_hieu")),
            normalize_inline(component.get("ten_cong_viec")),
            normalize_spaces(component.get("description")),
        ]
        return " | ".join(part for part in parts if part)
    return normalize_inline(component)

def build_components_suffix(components: List[Any]) -> str:
    component_texts = dedupe_preserve_order(
        [format_component_entry(component) for component in (components or []) if format_component_entry(component)]
    )
    if not component_texts:
        return ""
    return f'Gồm : "{ "; ".join(component_texts) }".'

def build_primary_match_summary(primary_match: Optional[dict]) -> str:
    pm = primary_match or {}
    item_name = normalize_vi_text(safe_str(pm.get("ten_cong_viec")))
    desc = normalize_vi_text(safe_str(pm.get("description")))
    parts: List[str] = []
    if item_name:
        parts.append(item_name.rstrip("."))
    if desc and normalize_basic_compare(desc) != normalize_basic_compare(item_name):
        parts.append(desc.rstrip("."))
    return ". ".join(part for part in parts if part).strip()

def ensure_components_description_format(text: str, primary_match: Optional[dict]) -> str:
    base_text = normalize_spaces(text)
    components = (primary_match or {}).get("components") or []
    suffix = build_components_suffix(components)
    if not suffix:
        return normalize_ai_description(base_text)

    cleaned = re.sub(r"\s*Gồm\s*:.*$", "", base_text, flags=re.IGNORECASE | re.DOTALL).strip()
    if len(cleaned) < 20:
        cleaned = build_primary_match_summary(primary_match) or cleaned
    if cleaned and cleaned[-1] not in ".!?":
        cleaned += "."
    final_text = f"{cleaned} {suffix}".strip() if cleaned else suffix
    return normalize_ai_description(final_text)

def remove_duplicate_tail_item(source_desc: str, item_name: str) -> str:
    """Loại bỏ tên hạng mục (cột B) trùng lặp trong mô tả (cột C).
    Xóa ở đầu, cuối, hoặc dòng riêng trùng khớp."""
    desc = normalize_spaces(source_desc)
    item = normalize_inline(item_name)
    if not desc or not item: return desc or ""
    item_cmp = normalize_basic_compare(item)
    if not item_cmp: return desc

    lines = [x for x in desc.split("\n") if x.strip()]

    # Xóa dòng trùng khớp hoàn toàn với tên B (đầu, cuối, giữa)
    filtered = [ln for ln in lines if normalize_basic_compare(ln) != item_cmp]

    # Nếu dòng đầu BẮT ĐẦU bằng tên B → cắt phần trùng, giữ phần sau
    if filtered:
        first = filtered[0]
        first_stripped = first.strip()
        item_stripped = item.strip()
        # So sánh case-insensitive: "1x15W LED batten ... 1x15W LED batten ..." → bỏ phần đầu
        if len(first_stripped) > len(item_stripped):
            first_lower = first_stripped.lower()
            item_lower = item_stripped.lower()
            if first_lower.startswith(item_lower):
                remainder = first_stripped[len(item_stripped):].lstrip(" .,;:-")
                if remainder:
                    filtered[0] = remainder[0].upper() + remainder[1:] if remainder else ""

    result = "\n".join(filtered).strip()
    return result if result else desc

def chunk_list(items: List[dict], size: int):
    for i in range(0, len(items), size):
        yield items[i : i + size]

def safe_json_loads(text: str):
    text = safe_str(text)
    if not text: return {}
    try: return json.loads(text)
    except Exception: pass
    match = re.search(r"\{.*\}", text, flags=re.DOTALL)
    if match:
        try: return json.loads(match.group(0))
        except Exception: pass
    return {}

def now_utc_iso() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds")

def sha1_text(text: str) -> str:
    return hashlib.sha1(safe_str(text).encode("utf-8")).hexdigest()

def _normalize_name_for_key(name: str) -> str:
    s = normalize_inline(name).rstrip(". ").lower()
    return re.sub(r"\s+", " ", s).strip()

def ensure_parent_dir(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)

def read_json_file(path: Path, default):
    try:
        if not path.exists(): return default
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception as exc:
        print(f"Không đọc được JSON {path}: {exc}")
        return default

def write_json_file_atomic(path: Path, data: Any) -> None:
    ensure_parent_dir(path)
    # Use a per-call temp path to avoid cross-thread/process collisions.
    tmp_path = path.parent / f".{path.name}.{os.getpid()}.{time.time_ns()}.tmp"
    try:
        with open(tmp_path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        try:
            os.replace(tmp_path, path)
        except PermissionError:
            try:
                if path.exists(): path.unlink()
                tmp_path.rename(path)
            except Exception:
                with open(path, "w", encoding="utf-8") as f:
                    json.dump(data, f, ensure_ascii=False, indent=2)
    finally:
        try:
            tmp_path.unlink(missing_ok=True)
        except Exception:
            pass


# =========================
# MULTI-INSTANCE LOAD HELPERS
# =========================
def _glob_instance_files(base_name: str, extension: str = ".json") -> List[Path]:
    return sorted(PERSIST_DIR.glob(f"{base_name}_*{extension}"))

def _legacy_file(base_name: str, extension: str = ".json") -> Path:
    return PERSIST_DIR / f"{base_name}{extension}"

def _my_instance_file(base_name: str, extension: str = ".json") -> Path:
    return PERSIST_DIR / f"{base_name}_{INSTANCE_ID}{extension}"

def _load_translation_from_file(path: Path, is_mine: bool) -> int:
    raw = read_json_file(path, [])
    if not isinstance(raw, list): return 0
    count = 0
    for e in raw:
        try:
            key = tuple(safe_str(x) for x in e.get("key", []))
            val = e.get("value", ["", ""])
            if len(key) == 4 and isinstance(val, list) and len(val) >= 2:
                nv, dv = safe_str(val[0]), safe_str(val[1])
                # Bỏ qua cache entry nếu kết quả vẫn còn tiếng nước ngoài
                if still_has_foreign(nv) or still_has_foreign(dv):
                    continue
                if key not in TRANSLATION_CACHE:
                    TRANSLATION_CACHE[key] = (nv, dv)
                    count += 1
                if is_mine:
                    _INST_TRANSLATION_KEYS.add(key)
        except Exception: continue
    return count

def _ai_desc_cache_entry_is_bad(val: str) -> bool:
    """Skip entries có Gồm: nhưng thiếu dấu ngoặc kép — sẽ được re-generate."""
    m = re.search(r'g[oồ]m\s*:', val, flags=re.IGNORECASE)
    if not m:
        return False
    return not val[m.end():].lstrip().startswith('"')

def _load_ai_desc_from_file(path: Path, is_mine: bool) -> int:
    raw = read_json_file(path, [])
    if not isinstance(raw, list): return 0
    count = 0
    for e in raw:
        try:
            key = tuple(safe_str(x) for x in e.get("key", []))
            val = safe_str(e.get("value", ""))
            if key and not _ai_desc_cache_entry_is_bad(val):
                if key not in AI_DESC_CACHE:
                    AI_DESC_CACHE[key] = val
                    count += 1
                if is_mine:
                    _INST_AI_DESC_KEYS.add(key)
        except Exception: continue
    return count

def _load_sim_audit_from_file(path: Path, is_mine: bool) -> int:
    raw = read_json_file(path, {})
    if not isinstance(raw, dict): return 0
    count = 0
    for k, v in raw.items():
        if not isinstance(v, dict): continue
        sk = safe_str(k)
        if sk not in SIM_AUDIT_CACHE:
            SIM_AUDIT_CACHE[sk] = {
                "best_code": safe_str(v.get("best_code")),
                "score":     float(v.get("score", 0.0) or 0.0),
                "label":     "true" if safe_str(v.get("label")).lower() == "true" else "false",
                "best_item": make_minimal_item(v.get("best_item")),
            }
            count += 1
        if is_mine:
            _INST_SIM_AUDIT_KEYS.add(sk)
    return count

def _load_output_cache_from_file(path: Path, is_mine: bool) -> int:
    raw = read_json_file(path, {})
    if not isinstance(raw, dict): return 0
    count = 0
    for k, v in raw.items():
        if k not in OUTPUT_CACHE:
            OUTPUT_CACHE[k] = v
            count += 1
        if is_mine:
            _INST_OUTPUT_CACHE_KEYS.add(k)
    return count

def _load_query_embed_from_npz(path: Path, is_mine: bool) -> int:
    if not path.exists(): return 0
    count = 0
    try:
        npz = np.load(path, allow_pickle=True)
        for k, v in zip(npz["keys"].tolist(), npz["vecs"]):
            sk = str(k)
            if sk not in QUERY_EMBED_CACHE:
                QUERY_EMBED_CACHE[sk] = v.astype(np.float32)
                count += 1
            if is_mine:
                _INST_QUERY_EMBED_KEYS.add(sk)
    except Exception as e:
        print(f"  query_embed npz lỗi ({path.name}): {e}")
    return count

def _load_query_embed_from_json(path: Path, is_mine: bool) -> int:
    raw = read_json_file(path, {})
    if not isinstance(raw, dict): return 0
    count = 0
    for k, v in raw.items():
        if isinstance(v, list) and k not in QUERY_EMBED_CACHE:
            QUERY_EMBED_CACHE[k] = np.array(v, dtype=np.float32)
            count += 1
        if is_mine:
            _INST_QUERY_EMBED_KEYS.add(k)
    return count

def _load_all_instances_cache(base_name: str, loader_fn, label: str) -> None:
    my_file   = _my_instance_file(base_name)
    legacy    = _legacy_file(base_name)
    instances = _glob_instance_files(base_name)
    total = 0
    if my_file.exists():
        total += loader_fn(my_file, is_mine=True)
    for f in instances:
        if f == my_file: continue
        total += loader_fn(f, is_mine=False)
    adopt_legacy = not my_file.exists()
    if legacy.exists():
        total += loader_fn(legacy, is_mine=adopt_legacy)
        if adopt_legacy:
            print(f"  [MIGRATE] {label}: adopt {legacy.name} → {my_file.name}")
    if total:
        print(f"  {label}: {total} entries loaded")


def make_minimal_item(item: Optional[dict]) -> Optional[dict]:
    if not item: return None
    return {
        "id": item.get("id"),
        "ma_hieu":       safe_str(item.get("ma_hieu")),
        "ten_cong_viec": safe_str(item.get("ten_cong_viec")),
        "description":   safe_str(item.get("description")),
        "embed_text":    safe_str(item.get("embed_text")),
        "source":        safe_str(item.get("source")),
        "learned_from":  safe_str(item.get("learned_from")),
        "learned_at":    safe_str(item.get("learned_at")),
    }


# =========================
# PERSISTENT STATE
# =========================
def load_persistent_state() -> None:
    global TRANSLATION_CACHE, AI_DESC_CACHE, SIM_AUDIT_CACHE, QUERY_EMBED_CACHE
    global LEARNED_KEYS_CACHE, LEARNED_ITEMS_RAM, _LEARNED_ITEMS_DIRTY
    global OUTPUT_CACHE, _OUTPUT_CACHE_DIRTY
    global _INST_TRANSLATION_KEYS, _INST_AI_DESC_KEYS, _INST_SIM_AUDIT_KEYS
    global _INST_OUTPUT_CACHE_KEYS, _INST_QUERY_EMBED_KEYS

    PERSIST_DIR.mkdir(parents=True, exist_ok=True)
    TRANSLATION_CACHE = {}; AI_DESC_CACHE = {}; SIM_AUDIT_CACHE = {}; QUERY_EMBED_CACHE = {}
    _INST_TRANSLATION_KEYS = set(); _INST_AI_DESC_KEYS = set()
    _INST_SIM_AUDIT_KEYS = set(); _INST_OUTPUT_CACHE_KEYS = set()
    _INST_QUERY_EMBED_KEYS = set()

    print(f"[LOAD] Instance '{INSTANCE_ID}' — đọc tất cả instance files + legacy...")

    _load_all_instances_cache("translation_cache", _load_translation_from_file, "Translation")
    _load_all_instances_cache("ai_desc_cache", _load_ai_desc_from_file, "AI desc")
    _load_all_instances_cache("sim_audit_cache", _load_sim_audit_from_file, "Sim audit")

    my_npz    = _my_instance_file("query_embed_cache", ".npz")
    all_npz   = _glob_instance_files("query_embed_cache", ".npz")
    legacy_npz= _legacy_file("query_embed_cache", ".npz")
    npz_adopt = not my_npz.exists()

    if my_npz.exists():
        _load_query_embed_from_npz(my_npz, is_mine=True)
    for f in all_npz:
        if f == my_npz: continue
        _load_query_embed_from_npz(f, is_mine=False)
    if legacy_npz.exists():
        _load_query_embed_from_npz(legacy_npz, is_mine=npz_adopt)

    if not QUERY_EMBED_CACHE:
        legacy_json = _legacy_file("query_embed_cache", ".json")
        if legacy_json.exists():
            _load_query_embed_from_json(legacy_json, is_mine=npz_adopt)

    _load_all_instances_cache("output_cache", _load_output_cache_from_file, "Output")

    existing_learned   = get_learned_knowledge_items()
    LEARNED_ITEMS_RAM  = list(existing_learned)
    _LEARNED_ITEMS_DIRTY = False
    LEARNED_KEYS_CACHE = {build_learned_item_key(it) for it in LEARNED_ITEMS_RAM}

    _OUTPUT_CACHE_DIRTY = False

    print(f"Translation cache: {len(TRANSLATION_CACHE)} | AI cache: {len(AI_DESC_CACHE)} "
          f"| Sim cache: {len(SIM_AUDIT_CACHE)} | Embed cache: {len(QUERY_EMBED_CACHE)} "
          f"| Learned: {len(LEARNED_KEYS_CACHE)} | Output cache: {len(OUTPUT_CACHE)}")

    _purge = [
        k for k, (nv, dv) in TRANSLATION_CACHE.items()
        if (not looks_like_vietnamese(k[0]) and not is_technical_code(k[0]) and k[0]
            and not looks_like_vietnamese(nv))
        or (not looks_like_vietnamese(k[1]) and k[1] and dv and not looks_like_vietnamese(dv))
    ]
    for k in _purge:
        del TRANSLATION_CACHE[k]
        _INST_TRANSLATION_KEYS.discard(k)
    if _purge:
        global _TRANSLATION_DIRTY
        _TRANSLATION_DIRTY = True
        print(f"  [PURGE] Xóa {len(_purge)} translation cache entries tiếng Anh sai.")


def save_persistent_state(force: bool = False) -> None:
    global _TRANSLATION_DIRTY, _AI_CACHE_DIRTY, _QUERY_EMBED_DIRTY
    global _LEARNED_ITEMS_DIRTY, _OUTPUT_CACHE_DIRTY

    if _TRANSLATION_DIRTY or force:
        write_json_file_atomic(TRANSLATION_CACHE_FILE, [
            {"key": list(k), "value": [safe_str(v[0]), safe_str(v[1])]}
            for k, v in TRANSLATION_CACHE.items()
            if k in _INST_TRANSLATION_KEYS
        ])
        _TRANSLATION_DIRTY = False

    if _AI_CACHE_DIRTY or force:
        write_json_file_atomic(AI_DESC_CACHE_FILE, [
            {"key": list(k), "value": safe_str(v)}
            for k, v in AI_DESC_CACHE.items()
            if k in _INST_AI_DESC_KEYS
        ])
        _AI_CACHE_DIRTY = False

    if _INST_SIM_AUDIT_KEYS and (force or True):
        sim_payload = {
            k: {
                "best_code": safe_str(v.get("best_code")),
                "score":     float(v.get("score", 0.0) or 0.0),
                "label":     "true" if safe_str(v.get("label")).lower() == "true" else "false",
                "best_item": make_minimal_item(v.get("best_item")),
            }
            for k, v in SIM_AUDIT_CACHE.items()
            if k in _INST_SIM_AUDIT_KEYS
        }
        write_json_file_atomic(SIM_AUDIT_CACHE_FILE, sim_payload)

    if _QUERY_EMBED_DIRTY or force:
        my_keys = [k for k in QUERY_EMBED_CACHE if k in _INST_QUERY_EMBED_KEYS]
        if my_keys:
            ks = my_keys
            # Luôn lưu dưới dạng numpy float32 để tương thích .npz
            vs = np.array([
                QUERY_EMBED_CACHE[k].cpu().numpy()
                if (HAS_TORCH and hasattr(QUERY_EMBED_CACHE[k], "cpu"))
                else np.asarray(QUERY_EMBED_CACHE[k], dtype=np.float32)
                for k in ks
            ], dtype=np.float32)
            ensure_parent_dir(QUERY_EMBED_CACHE_NPZ_FILE)
            np.savez_compressed(QUERY_EMBED_CACHE_NPZ_FILE,
                                keys=np.array(ks, dtype=object), vecs=vs)
        _QUERY_EMBED_DIRTY = False

    if _LEARNED_ITEMS_DIRTY or force:
        my_learned = [it for it in LEARNED_ITEMS_RAM
                      if it.get("instance_id", INSTANCE_ID) == INSTANCE_ID]
        save_learned_knowledge_items(my_learned)
        _LEARNED_ITEMS_DIRTY = False

    if _OUTPUT_CACHE_DIRTY or force:
        my_output = {k: v for k, v in OUTPUT_CACHE.items()
                     if k in _INST_OUTPUT_CACHE_KEYS}
        write_json_file_atomic(OUTPUT_CACHE_FILE, my_output)
        _OUTPUT_CACHE_DIRTY = False


# =========================
# VI HELPERS
# =========================
VIETNAMESE_CHAR_RE = re.compile(
    r"[àáảãạăắằẳẵặâấầẩẫậđèéẻẽẹêếềểễệìíỉĩịòóỏõọôốồổỗộơớờởỡợùúủũụưứừửữựỳýỷỹỵ]",
    flags=re.IGNORECASE,
)
COMMON_VI_MARKERS = [
    "cung cấp","lắp đặt","thiết bị","đường ống","máng cáp","thang cáp",
    "tủ điện","bản vẽ","yêu cầu kỹ thuật","phụ kiện","ống","van","quạt",
    "cửa gió","cáp","hoàn chỉnh",
]

def looks_like_vietnamese(text: str) -> bool:
    text = normalize_spaces(text)
    if not text: return False
    if VIETNAMESE_CHAR_RE.search(text): return True
    low   = text.lower()
    score = sum(1 for m in COMMON_VI_MARKERS if m in low)
    return score >= 2

_ENGLISH_WORD_RE = re.compile(
    r"\b(supply|install|provide|including|complete|unit|panel|board|cable|pipe|"
    r"valve|pump|fan|duct|grille|diffuser|fitting|accessory|accessories|"
    r"accordance|drawing|specification|general|meter|system|equipment|"
    r"and|or|with|for|the|to|of|in|at|by|from|per|each)\b",
    flags=re.IGNORECASE,
)

def looks_like_english(text: str) -> bool:
    text = normalize_spaces(text)
    if not text: return False
    if looks_like_vietnamese(text): return False
    return bool(_ENGLISH_WORD_RE.search(text))

_LATIN_WORD_RE = re.compile(r"\b[A-Za-z]{4,}\b")
_VIET_DIACRITIC_WORD_RE = re.compile(
    r"\b\w*[àáảãạăắằẳẵặâấầẩẫậđèéẻẽẹêếềểễệìíỉĩịòóỏõọôốồổỗộơớờởỡợùúủũụưứừửữựỳýỷỹỵ]\w*\b",
    flags=re.IGNORECASE,
)
# CJK: Chinese, Japanese (Hiragana/Katakana), Korean Hangul
_CJK_CHAR_RE = re.compile(
    r"[\u4e00-\u9fff\u3400-\u4dbf\uf900-\ufaff"  # Chinese CJK
    r"\u3040-\u309f\u30a0-\u30ff"                # Japanese Hiragana/Katakana
    r"\uac00-\ud7af\u1100-\u11ff]"               # Korean Hangul
)

def is_foreign_short(text: str) -> bool:
    """Cho text ngắn (item name): ngoại ngữ nếu:
    - Có ký tự CJK (Trung/Nhật/Hàn), HOẶC
    - Không có dấu tiếng Việt VÀ có >=1 từ Latin"""
    text = normalize_spaces(text)
    if not text or len(text) < 2: return False
    if _CJK_CHAR_RE.search(text): return True
    if _VIET_DIACRITIC_WORD_RE.search(text): return False
    return bool(_LATIN_WORD_RE.search(text))

def still_has_foreign(text: str) -> bool:
    """Check text có chủ yếu là ngoại ngữ không.
    Trả True nếu:
    - Có ký tự CJK (Trung/Nhật/Hàn), HOẶC
    - Số từ Latin >=3 VÀ >= từ có dấu tiếng Việt"""
    text = normalize_spaces(text)
    if not text or len(text) < 2: return False
    # CJK characters → chắc chắn là ngoại ngữ
    if _CJK_CHAR_RE.search(text): return True
    if len(text) < 10: return False
    latin_words = _LATIN_WORD_RE.findall(text)
    viet_words  = _VIET_DIACRITIC_WORD_RE.findall(text)
    return len(latin_words) >= 3 and len(latin_words) >= len(viet_words)

def is_technical_code(text: str) -> bool:
    t = normalize_inline(text)
    if not t: return True
    return bool(re.fullmatch(r'[A-Z0-9\-\/\.\(\)\+xX\s,;:]+', t))

_ROMAN_NUMERAL_RE = re.compile(
    r'^M{0,4}(CM|CD|D?C{0,3})(XC|XL|L?X{0,3})(IX|IV|V?I{0,3})$', re.IGNORECASE
)

def is_roman_numeral_code(code: str) -> bool:
    """Trả True nếu code là số La Mã thuần (không dấu chấm) → phân mục cha, không sinh mô tả."""
    c = code.strip()
    if not c or "." in c:
        return False
    m = _ROMAN_NUMERAL_RE.match(c)
    return bool(m) and bool(m.group(0))  # loại chuỗi rỗng khớp regex

COMMON_ITEM_REPLACEMENTS = [
    (r"\bmain rmu panel\b","Tủ RMU chính"),(r"\brmu panel\b","Tủ RMU"),
    (r"\bdistribution board\b","Tủ điện phân phối"),(r"\bswitchboard\b","Tủ điện"),
    (r"\bcable ladder\b","Thang cáp"),(r"\bcable tray\b","Máng cáp"),
    (r"\btrunking\b","Máng điện"),(r"\bconduit\b","Ống luồn dây"),
    (r"\bpipe sleeve\b","Ống sleeve"),(r"\bgate valve\b","Van cổng"),
    (r"\bcheck valve\b","Van một chiều"),(r"\bbutterfly valve\b","Van bướm"),
    (r"\bball valve\b","Van bi"),(r"\bair grille\b","Cửa gió"),
    (r"\bair grill\b","Cửa gió"),(r"\blouver\b","Cửa chớp gió"),
    (r"\bdiffuser\b","Miệng gió khuếch tán"),(r"\bfan\b","Quạt"),
]
COMMON_DESC_REPLACEMENTS = [
    (r"Supply and install","Cung cấp và lắp đặt"),
    (r"Supply & Install","Cung cấp và lắp đặt"),
    (r"Supply and installation","Cung cấp và lắp đặt"),
    (r"including all necessary fittings and accessories","bao gồm đầy đủ phụ kiện và vật tư cần thiết"),
    (r"including all necessary items and accessories","bao gồm đầy đủ vật tư và phụ kiện cần thiết"),
    (r"all in accordance with drawings and specifications","theo bản vẽ và yêu cầu kỹ thuật"),
    (r"all following drawings and specifications","theo bản vẽ và yêu cầu kỹ thuật"),
    (r"complete","hoàn chỉnh"),
    (r"Gate valve","Van cổng"),(r"Check valve","Van một chiều"),
    (r"air grill","cửa gió"),(r"louver","cửa chớp gió"),
    (r"diffuser","miệng gió khuếch tán"),(r"air valve","van gió"),
    (r"cables","cáp điện"),(r"cable trays","máng cáp"),
    (r"trunking","máng điện"),(r"conduits?","ống luồn dây"),
    (r"floor drain","phễu thu sàn"),(r"ventcowl","chụp thông hơi"),
    (r"c/w","bao gồm"),
]

def normalize_vi_text(text: str) -> str:
    text = normalize_spaces(text)
    if not text: return ""
    replacements = {
        r"\bsq\.mm\b":"mm²",r"\bsqmm\b":"mm²",r"\bmm2\b":"mm²",
        r"\bkv\b":"kV",r"\bdn\s*([0-9]+)\b":r"DN\1",
        r"\bxlpe\b":"XLPE",r"\bpvc\b":"PVC",r"\bcu\b":"Cu",
        r"\bupvc\b":"uPVC",r"\bhdpe\b":"HDPE",r"\brmu\b":"RMU",
        r"\bmsb\b":"MSB",r"\bmdb\b":"MDB",r"\bdb\b":"DB",
        r"\bahu\b":"AHU",r"\bfcu\b":"FCU",
    }
    for pat, repl in replacements.items():
        text = re.sub(pat, repl, text, flags=re.IGNORECASE)
    text = text.replace(" ,",",").replace(" .",".").replace(" ;",";").replace(" :",":")
    lines = []
    for line in text.split("\n"):
        line = line.strip(" \t-;,")
        if not line: continue
        line = line[0].upper() + line[1:] if line else line
        lines.append(line)
    return "\n".join(lines).strip()

def fallback_translate_item_name(source_item_name: str, source_desc: str = "") -> str:
    text = normalize_inline(source_item_name)
    if not text: return ""
    if looks_like_vietnamese(text): return normalize_vi_text(text)
    out = text
    for pat, repl in COMMON_ITEM_REPLACEMENTS:
        out = re.sub(pat, repl, out, flags=re.IGNORECASE)
    return normalize_vi_text(out)

def fallback_translate_desc(source_desc: str, source_item_name: str = "") -> str:
    text = remove_duplicate_tail_item(source_desc, source_item_name)
    if not text: return ""
    if looks_like_vietnamese(text): return normalize_vi_text(text)
    out = text
    for pat, repl in COMMON_DESC_REPLACEMENTS:
        out = re.sub(pat, repl, out, flags=re.IGNORECASE)
    out = re.sub(r"\s*;\s*","; ",out); out = re.sub(r"\s*,\s*",", ",out)
    out = re.sub(r"\s+"," ",out); out = out.replace(" & "," và ")
    return normalize_vi_text(out)


# =========================
# XLS / XLSX HELPERS
# =========================
def find_soffice_executable():
    candidates = [
        shutil.which("soffice"), shutil.which("libreoffice"),
        r"C:\Program Files\LibreOffice\program\soffice.exe",
        r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
    ]
    for p in candidates:
        if p and os.path.exists(p): return p
    return None

def convert_xls_to_xlsx(input_path: str) -> str:
    soffice = find_soffice_executable()
    if not soffice:
        raise RuntimeError("Không tìm thấy LibreOffice. Lưu file thành .xlsx trước.")
    input_path = str(Path(input_path).resolve())
    out_dir    = tempfile.mkdtemp(prefix="xls_to_xlsx_")
    result     = subprocess.run(
        [soffice,"--headless","--convert-to","xlsx","--outdir",out_dir,input_path],
        capture_output=True, text=True,
    )
    if result.returncode != 0:
        raise RuntimeError(f"Chuyển .xls thất bại.\n{result.stderr}")
    converted = os.path.join(out_dir, f"{Path(input_path).stem}.xlsx")
    if not os.path.exists(converted):
        files = list(Path(out_dir).glob("*.xlsx"))
        if not files: raise RuntimeError("Không tìm thấy file .xlsx sau chuyển đổi.")
        converted = str(files[0])
    return converted

def get_readable_workbook_path(input_file: str) -> str:
    ext = Path(input_file).suffix.lower()
    if ext in {".xlsx",".xlsm"}: return input_file
    if ext == ".xls":
        print("Đang chuyển .xls sang .xlsx ...")
        return convert_xls_to_xlsx(input_file)
    raise ValueError("Chỉ hỗ trợ .xls, .xlsx hoặc .xlsm")


# =========================
# EXCEL HELPERS
# =========================
def ensure_header(ws, header_name: str, preferred_col: Optional[int] = None) -> int:
    header_map = {}
    for col in range(1, ws.max_column + 1):
        h = normalize_inline(get_cell_value(ws, HEADER_ROW, col))
        if h: header_map[h] = col
    if header_name in header_map: return header_map[header_name]
    if preferred_col is None: preferred_col = ws.max_column + 1
    ws.cell(row=HEADER_ROW, column=preferred_col, value=header_name)
    ws.cell(row=HEADER_ROW, column=preferred_col).alignment = Alignment(wrap_text=True, vertical="top")
    return preferred_col

def get_column_map(ws, read_only: bool = False) -> Dict[str, Optional[int]]:
    header_map: Dict[str, int] = {}
    if read_only:
        for row in ws.iter_rows(min_row=HEADER_ROW, max_row=HEADER_ROW, values_only=False):
            for cell in row:
                h = normalize_inline(safe_str(cell.value))
                if h: header_map[h] = cell.column
    else:
        for col in range(1, ws.max_column + 1):
            h = normalize_inline(get_cell_value(ws, HEADER_ROW, col))
            if h: header_map[h] = col
    ai_col = header_map.get(HEADER_AI_DESC, OUTPUT_AI_COL_FALLBACK)
    if not read_only:
        ai_col = ensure_header(ws, HEADER_AI_DESC, ai_col)
    nhan_hieu_col = None
    _nh_headers_lower = {normalize_inline(h).lower() for h in NHAN_HIEU_HEADERS}
    for h_key, col in header_map.items():
        if h_key.lower() in _nh_headers_lower:
            nhan_hieu_col = col
            break
    return {
        "item_code":            header_map.get(HEADER_ITEM_CODE, header_map.get(HEADER_PHAN_MUC, ITEM_CODE_COL_FALLBACK)),
        "item_name":            header_map.get(HEADER_ITEM_NAME,             ITEM_NAME_COL_FALLBACK),
        "source_desc":          header_map.get(HEADER_SOURCE_DESC,           SOURCE_DESC_COL_FALLBACK),
        "translated_item_name": header_map.get(HEADER_TRANSLATED_ITEM_NAME,  TRANSLATED_ITEM_NAME_COL_FALLBACK),
        "translated_desc":      header_map.get(HEADER_TRANSLATED_DESC,       TRANSLATED_DESC_COL_FALLBACK),
        "ai_desc":              ai_col,
        "nhan_hieu":            nhan_hieu_col,
    }

def set_col_width_if_smaller(ws, col_idx: int, width: float):
    col_letter = get_column_letter(col_idx)
    current    = ws.column_dimensions[col_letter].width
    if current is None or current < width:
        ws.column_dimensions[col_letter].width = width

def setup_columns(ws, col_map: Dict[str, Optional[int]]):
    if col_map["translated_item_name"]: set_col_width_if_smaller(ws, col_map["translated_item_name"], 38)
    if col_map["translated_desc"]:      set_col_width_if_smaller(ws, col_map["translated_desc"],      58)
    if col_map["ai_desc"]:              set_col_width_if_smaller(ws, col_map["ai_desc"],              70)
    for key in ["translated_item_name","translated_desc","ai_desc"]:
        col = col_map.get(key)
        if col: ws.cell(row=HEADER_ROW, column=col).alignment = Alignment(wrap_text=True, vertical="top")

def get_cell_value(ws, row_idx: int, col_idx: int):
    try:
        cell = ws.cell(row=row_idx, column=col_idx)
        from openpyxl.cell.cell import MergedCell
        if isinstance(cell, MergedCell): return None
        return cell.value
    except Exception: return None

def write_cell(ws, row_idx: int, col_idx: int, value: Any):
    ws.cell(row=row_idx, column=col_idx, value=value)
    ws.cell(row=row_idx, column=col_idx).alignment = Alignment(wrap_text=True, vertical="top")

def estimate_row_height(*values: Any) -> int:
    line_count = 1
    for value in values:
        if value is None or value == "": continue
        lines      = str(value).split("\n")
        line_count += sum(max(1, (len(line)//50)+1) for line in lines)
    return int(max(ROW_HEIGHT_MIN, min(ROW_HEIGHT_MAX, line_count * 14)))


# =========================
# KNOWLEDGE / EMBEDDING  (GPU-accelerated)
# =========================
MA_HIEU_TOKEN_RE = re.compile(
    r"\b(?:[A-Za-z]{1,20}[/-][A-Za-z0-9+.,xX:-]+|\d{2,4}x\d{2,4}x\d(?:[.,]\d+)?-[A-Za-z0-9-]+)\b"
)
PANEL_PREFIX_MAP = {
    "DB":"Tủ điện phân phối","MDB":"Tủ điện phân phối chính",
    "SMDB":"Tủ điện phân phối phụ","SDB":"Tủ điện phân phối phụ",
    "MSB":"Tủ điện tổng","MCC":"Tủ điều khiển động cơ",
    "ACP":"Tủ điện điều hòa","ATS":"Tủ chuyển nguồn tự động",
    "UPS":"Bộ lưu điện UPS","LP":"Tủ chiếu sáng",
    "PP":"Tủ điện động lực","CP":"Tủ điều khiển",
    "FACP":"Tủ báo cháy trung tâm","FAP":"Tủ báo cháy",
    "EMDB":"Tủ điện khẩn cấp","EPP":"Tủ điện động lực khẩn cấp",
    "ELP":"Tủ điện chiếu sáng khẩn cấp","FCP":"Tủ điều khiển quạt",
    "RMU":"Tủ RMU",
}

def normalize_ma_hieu(text: str) -> str:
    s = safe_str(text).upper()
    if not s: return ""
    s = s.replace("×","X").replace("*","X").replace("_","-")
    s = s.replace(" ","").replace("–","-").replace("—","-")
    s = re.sub(r"-{2,}","-",s)
    return s.strip("-")

def extract_ma_hieu_from_text(*texts: str) -> str:
    candidates = []
    for text in texts:
        raw = safe_str(text)
        if not raw: continue
        for m in MA_HIEU_TOKEN_RE.finditer(raw):
            token      = normalize_ma_hieu(m.group(0))
            digit_count= sum(ch.isdigit() for ch in token)
            if digit_count >= 2 and ("-" in token or "/" in token or "X" in token):
                candidates.append(token)
    if not candidates: return ""
    return sorted(set(candidates), key=lambda x: (-len(x), x))[0]

def split_code_tokens(code: str) -> List[str]:
    code = normalize_ma_hieu(code)
    if not code: return []
    return [x for x in code.split("-") if x]

def get_learned_knowledge_items() -> List[dict]:
    all_items: List[dict] = []
    for f in _glob_instance_files("learned_knowledge"):
        data = read_json_file(f, [])
        if isinstance(data, list):
            all_items.extend(data)
    legacy = _legacy_file("learned_knowledge")
    if legacy.exists():
        data = read_json_file(legacy, [])
        if isinstance(data, list):
            all_items.extend(data)
    seen: set = set()
    deduped: List[dict] = []
    for item in all_items:
        if not isinstance(item, dict): continue
        key = build_learned_item_key(item)
        if key not in seen:
            seen.add(key)
            deduped.append(item)
    return deduped

def save_learned_knowledge_items(items: List[dict]) -> None:
    write_json_file_atomic(LEARNED_KNOWLEDGE_FILE, items)

def build_learned_item_key(item: dict) -> str:
    norm_code = normalize_ma_hieu(item.get("ma_hieu",""))
    if norm_code: return f"code::{norm_code}"
    name = _normalize_name_for_key(item.get("ten_cong_viec",""))
    return f"name::{sha1_text(name[:80])}"

def merge_knowledge_items(base_items: List[dict], learned_items: List[dict]) -> List[dict]:
    merged, seen = [], set()
    for src, source_name in [(base_items,"base"),(learned_items,"learned")]:
        for item in src:
            if not isinstance(item, dict): continue
            norm_code  = normalize_ma_hieu(item.get("ma_hieu",""))
            key        = build_learned_item_key(item)
            dedupe_key = f"code::{norm_code}" if norm_code else key
            if dedupe_key in seen: continue
            seen.add(dedupe_key)
            new_item = dict(item)
            new_item["source"] = safe_str(new_item.get("source")) or source_name
            merged.append(new_item)
    return merged


def _embed_items_with_cache(
    items_to_embed: List[dict],
    name_texts: List[str],
    meta_file: Path,
    vec_file: Path,
    label: str = "items",
    source_file: Optional[Path] = None,
) -> np.ndarray:
    """Embed danh sách items trên GPU — cache bằng .npz numpy (tương thích)."""
    norm_texts = [normalize_inline(x) for x in name_texts]
    n          = len(norm_texts)
    if n == 0:
        return np.zeros((0, 384), dtype=np.float32)

    mtime = 0
    if source_file and source_file.exists():
        mtime = int(source_file.stat().st_mtime)
    else:
        sample = items_to_embed[:5] + items_to_embed[-3:] if n > 8 else items_to_embed
        mtime  = int(sha1_text("||".join(
            normalize_inline(it.get("ten_cong_viec",""))[:20] for it in sample
        ))[:8], 16)

    signature = sha1_text(f"mtime:{mtime}:n:{n}:{EMBED_MODEL_NAME}:{DEVICE}")

    meta = read_json_file(meta_file, {})
    if (isinstance(meta, dict)
            and safe_str(meta.get("signature")) == signature
            and safe_str(meta.get("model_name")) == EMBED_MODEL_NAME
            and vec_file.exists()):
        try:
            data    = np.load(vec_file, allow_pickle=True)
            vecs_c  = data["vectors"].astype(np.float32)
            if vecs_c.shape[0] == n:
                print(f"  Cache OK ({label}): {n} vectors")
                return vecs_c
        except Exception as e:
            print(f"  Cache load lỗi ({label}): {e}")

    print(f"  Embedding {n} {label} trên {DEVICE} (batch={EMBED_BATCH_SIZE}) ...")
    model = get_embed_model()
    vecs  = model.encode(
        norm_texts,
        convert_to_numpy=True,
        normalize_embeddings=True,
        show_progress_bar=True,
        batch_size=EMBED_BATCH_SIZE,   # ← GPU batch size lớn hơn nhiều
    ).astype(np.float32)

    write_json_file_atomic(meta_file, {
        "signature":  signature,
        "model_name": EMBED_MODEL_NAME,
        "device":     DEVICE,
        "count":      n,
        "mtime":      mtime,
        "saved_at":   now_utc_iso(),
    })
    ensure_parent_dir(vec_file)
    np.savez_compressed(vec_file, vectors=vecs)
    return vecs


def _get_gpu_sim_matrix(knowledge: dict):
    """Trả về similarity matrix trên GPU (torch.Tensor).
    Tự rebuild khi knowledge signature thay đổi.
    Nếu torch không có → trả về numpy matrix (fallback).
    """
    global _GPU_SIM_MATRIX, _GPU_SIM_MATRIX_SIG

    sig = knowledge.get("signature", "")
    if _GPU_SIM_MATRIX is not None and _GPU_SIM_MATRIX_SIG == sig:
        return _GPU_SIM_MATRIX

    matrix_np = knowledge["similarity_pool_matrix"]  # numpy float32

    if HAS_TORCH and DEVICE != "cpu":
        t = torch.from_numpy(matrix_np).to(DEVICE)
        # FP16 tiết kiệm VRAM và nhanh hơn trên Tensor Cores
        if DEVICE == "cuda":
            t = t.half()
        _GPU_SIM_MATRIX     = t
        _GPU_SIM_MATRIX_SIG = sig
        print(f"  [GPU] Similarity matrix lên {DEVICE}: "
              f"{matrix_np.shape} dtype={t.dtype} "
              f"mem≈{t.numel()*t.element_size()/1e6:.1f}MB")
    else:
        # CPU fallback: giữ nguyên numpy
        _GPU_SIM_MATRIX     = matrix_np
        _GPU_SIM_MATRIX_SIG = sig

    return _GPU_SIM_MATRIX


def load_knowledge_index() -> dict:
    global KNOWLEDGE_INDEX
    if KNOWLEDGE_INDEX is not None: return KNOWLEDGE_INDEX

    if not os.path.exists(KNOWLEDGE_JSON_FILE):
        raise FileNotFoundError(f"Không tìm thấy knowledge json: {KNOWLEDGE_JSON_FILE}")

    with open(KNOWLEDGE_JSON_FILE, "r", encoding="utf-8") as f:
        base_raw = json.load(f)
    # Gộp thêm boq_merged_local nếu có
    if os.path.exists(KNOWLEDGE_LOCAL_FILE):
        with open(KNOWLEDGE_LOCAL_FILE, "r", encoding="utf-8") as f:
            local_raw = json.load(f)
        # Loại bỏ items trùng id
        existing_ids = {item.get("id") for item in base_raw}
        for item in local_raw:
            if item.get("id") not in existing_ids:
                base_raw.append(item)
        print(f"[Knowledge] Merged local: +{len(base_raw) - len(existing_ids)} items from {KNOWLEDGE_LOCAL_FILE}")
    boq_component_lookup = load_boq_component_lookup()

    learned_raw = get_learned_knowledge_items()

    def _parse_items(raw_list: List[dict], use_boq_components: bool = False) -> Tuple[List[dict], List[str]]:
        items, texts = [], []
        for raw in raw_list:
            ma_hieu    = safe_str(raw.get("ma_hieu"))
            ten        = safe_str(raw.get("ten_cong_viec"))
            desc       = safe_str(raw.get("description"))
            embed_text = safe_str(raw.get("embed_text"))
            norm_ma    = normalize_ma_hieu(ma_hieu)
            parts      = [p for p in [normalize_inline(ten), normalize_spaces(desc), norm_ma] if p]
            combined   = " | ".join(parts)
            name_for   = combined or normalize_inline(embed_text) or norm_ma
            if not norm_ma and not name_for: continue
            items.append({
                "id": raw.get("id"), "ma_hieu": ma_hieu, "norm_ma_hieu": norm_ma,
                "ten_cong_viec": ten, "description": desc, "embed_text": embed_text,
                "components": resolve_components_for_item(raw, boq_component_lookup) if use_boq_components else (raw.get("components") or []),
                "source":        safe_str(raw.get("source")),
                "learned_from":  safe_str(raw.get("learned_from")),
                "learned_at":    safe_str(raw.get("learned_at")),
            })
            texts.append(name_for)
        return items, texts

    base_items,    base_texts    = _parse_items(base_raw, use_boq_components=True)
    learned_items, learned_texts = _parse_items(learned_raw)

    print(f"[Knowledge] base={len(base_items)}, learned={len(learned_items)}")
    base_vecs = _embed_items_with_cache(
        base_items, base_texts,
        KNOWLEDGE_CACHE_META_FILE, KNOWLEDGE_CACHE_VEC_FILE,
        label="base knowledge items",
        source_file=Path(KNOWLEDGE_JSON_FILE),
    )

    _learned_files = _glob_instance_files("learned_knowledge") + [_legacy_file("learned_knowledge")]
    _latest_learned = max(
        (f for f in _learned_files if f.exists()),
        key=lambda f: f.stat().st_mtime,
        default=LEARNED_KNOWLEDGE_FILE,
    )
    if learned_items:
        learned_vecs = _embed_items_with_cache(
            learned_items, learned_texts,
            LEARNED_CACHE_META_FILE, LEARNED_CACHE_VEC_FILE,
            label="learned items",
            source_file=_latest_learned,
        )
    else:
        learned_vecs = np.zeros((0, base_vecs.shape[1] if base_vecs.size else 384), dtype=np.float32)

    for item, vec in zip(base_items, base_vecs):
        item["name_vec"] = vec
    for item, vec in zip(learned_items, learned_vecs):
        item["name_vec"] = vec

    all_items: List[dict] = []
    by_code:   Dict[str, dict] = {}
    by_exact_name: Dict[str, dict] = {}
    seen_keys: set = set()
    for item in base_items + learned_items:
        key = build_learned_item_key(item)
        if key in seen_keys: continue
        seen_keys.add(key)
        all_items.append(item)
        if item["norm_ma_hieu"] and item["norm_ma_hieu"] not in by_code:
            by_code[item["norm_ma_hieu"]] = item
        exact_name_key = _exact_name_key(item.get("ten_cong_viec", ""))
        if exact_name_key:
            existing = by_exact_name.get(exact_name_key)
            if existing is None:
                by_exact_name[exact_name_key] = item
            elif not (existing.get("components") or []) and (item.get("components") or []):
                by_exact_name[exact_name_key] = item

    sim_matrix = np.vstack([x["name_vec"] for x in all_items]).astype(np.float32) if all_items \
                 else np.zeros((0, 384), dtype=np.float32)

    KNOWLEDGE_INDEX = {
        "items": all_items, "items_with_code": list(by_code.values()), "by_code": by_code,
        "by_exact_name": by_exact_name,
        "similarity_pool_items": all_items, "similarity_pool_matrix": sim_matrix,
        "signature": sha1_text(f"{len(base_items)}:{len(learned_items)}"),
    }
    print(f"  Knowledge index sẵn sàng: {len(all_items)} items "
          f"(base:{len(base_items)}, learned:{len(learned_items)})")

    # Preload GPU matrix ngay khi build xong
    _get_gpu_sim_matrix(KNOWLEDGE_INDEX)
    return KNOWLEDGE_INDEX


def embed_texts(texts: List[str], persist_query_cache: bool = True) -> np.ndarray:
    """Embed texts trên GPU, trả về numpy float32 (tương thích với code cũ).
    Cache lưu dưới dạng numpy để tương thích .npz.
    """
    global _QUERY_EMBED_DIRTY
    if not texts: return np.zeros((0,1),dtype=np.float32)

    normalized        = [normalize_inline(x) for x in texts]
    result_vecs       = [None] * len(normalized)
    missing_positions = []
    missing_texts_ne  = []
    empty_positions   = []
    dim               = 384

    for i, text in enumerate(normalized):
        if not text:
            empty_positions.append(i); continue
        cached = QUERY_EMBED_CACHE.get(text)
        if cached is not None:
            v = cached if isinstance(cached, np.ndarray) else np.array(cached, dtype=np.float32)
            result_vecs[i] = v; dim = v.shape[0]
        else:
            missing_positions.append(i); missing_texts_ne.append(text)

    if not missing_texts_ne:
        for pos in empty_positions: result_vecs[pos] = np.zeros(dim, dtype=np.float32)
        return np.vstack(result_vecs).astype(np.float32)

    model    = get_embed_model()
    new_vecs = model.encode(
        missing_texts_ne,
        convert_to_numpy=True,
        normalize_embeddings=True,
        show_progress_bar=False,
        batch_size=EMBED_BATCH_SIZE,   # ← GPU batch size
    ).astype(np.float32)
    dim = new_vecs.shape[1]

    for pos, text, vec in zip(missing_positions, missing_texts_ne, new_vecs):
        result_vecs[pos] = vec
        if persist_query_cache:
            # Luôn lưu numpy trong RAM cache (không lưu tensor GPU để tránh OOM)
            QUERY_EMBED_CACHE[text] = vec
            _INST_QUERY_EMBED_KEYS.add(text)
            _QUERY_EMBED_DIRTY = True

    for pos in empty_positions:
        result_vecs[pos] = np.zeros(dim, dtype=np.float32)

    return np.vstack(result_vecs).astype(np.float32)


def _cosine_sim_matrix(query_vecs: np.ndarray, knowledge: dict) -> np.ndarray:
    """Tính cosine similarity trên GPU nếu có torch+CUDA/MPS, fallback numpy.

    - Input:  query_vecs  numpy (Q, D) float32
    - Output: score_matrix numpy (Q, K) float32  ← luôn trả về numpy để code sau tương thích
    """
    gpu_matrix = _get_gpu_sim_matrix(knowledge)

    if HAS_TORCH and DEVICE != "cpu":
        # Chuyển query lên GPU
        qvecs_t = torch.from_numpy(query_vecs).to(DEVICE)
        if DEVICE == "cuda":
            qvecs_t = qvecs_t.half()      # FP16 cho phù hợp với matrix

        # matmul trên GPU: (Q, D) x (D, K) → (Q, K)
        with torch.no_grad():
            scores_t = torch.mm(qvecs_t, gpu_matrix.T)

        # Chuyển về CPU numpy ngay lập tức để giải phóng VRAM
        return scores_t.float().cpu().numpy()
    else:
        # CPU numpy fallback
        return np.matmul(query_vecs, gpu_matrix.T)


def build_match_payload(item: dict, score: float, match_type: str) -> dict:
    return {
        "matched": True, "match_type": match_type, "score": round(float(score),6),
        "ma_hieu":       safe_str(item.get("ma_hieu")),
        "ten_cong_viec": safe_str(item.get("ten_cong_viec")),
        "description":   safe_str(item.get("description")),
        "components":    item.get("components") or [],
        "embed_text":    safe_str(item.get("embed_text")),
        "source":        safe_str(item.get("source")),
        "learned_from":  safe_str(item.get("learned_from")),
        "learned_at":    safe_str(item.get("learned_at")),
        "best_item": item,
    }

def get_output_cache_key(source_item_name: str, source_desc_cleaned: str) -> str:
    payload = f"{normalize_inline(source_item_name)}||{normalize_spaces(source_desc_cleaned)}"
    return sha1_text(payload[:300])

def get_sim_cache_key(source_item_name: str, source_desc_cleaned: str) -> str:
    payload = f"{normalize_inline(source_item_name)}||{normalize_spaces(source_desc_cleaned)}"
    return sha1_text(payload[:300])

def _exact_name_key(text: str) -> str:
    return normalize_inline(text).lower()


def find_best_similarity_code_by_name_batch(row_to_name: Dict[int, str]) -> Dict[int, dict]:
    """GPU-accelerated batch similarity search."""
    knowledge  = load_knowledge_index()
    pool       = knowledge["similarity_pool_items"]
    results: Dict[int, dict] = {}

    if not pool or knowledge["similarity_pool_matrix"].shape[0] == 0:
        for row_idx in row_to_name: results[row_idx] = {"best_code":"","score":0.0,"label":"false","best_item":None}
        return results

    row_to_key: Dict[int,str] = {}; unique_queries: List[str] = []; query_pos: Dict[str,int] = {}

    for row_idx, name in row_to_name.items():
        key = normalize_inline(name); row_to_key[row_idx] = key
        if not key:
            results[row_idx] = {"best_code":"","score":0.0,"label":"false","best_item":None}; continue
        if key not in query_pos:
            query_pos[key] = len(unique_queries); unique_queries.append(key)

    if unique_queries:
        qvecs        = embed_texts(unique_queries, persist_query_cache=True)

        # ── GPU cosine similarity ─────────────────────────────────────────────
        score_matrix = _cosine_sim_matrix(qvecs, knowledge)   # numpy (Q, K)

        TOP_K        = 5
        query_results: Dict[str, dict] = {}

        for q, pos in query_pos.items():
            row_scores  = score_matrix[pos]
            k           = min(TOP_K, len(pool))
            topk_idx    = np.argpartition(row_scores,-k)[-k:]
            topk_idx    = topk_idx[np.argsort(-row_scores[topk_idx])]
            q_tokens    = set(re.sub(r"[^\w]"," ",q.lower()).split())
            best_item   = pool[int(topk_idx[0])]
            best_cos    = float(row_scores[topk_idx[0]])
            best_combined = best_cos * 0.8

            for idx in topk_idx:
                item = pool[int(idx)]; cos = float(row_scores[idx])
                if best_cos - cos > 0.05: break
                item_text  = " ".join([safe_str(item.get("ten_cong_viec","")),
                                       safe_str(item.get("description","")),
                                       safe_str(item.get("ma_hieu",""))]).lower()
                item_tokens= set(re.sub(r"[^\w]"," ",item_text).split())
                overlap    = len(q_tokens & item_tokens) / max(len(q_tokens),1)
                combined   = cos * 0.8 + overlap * 0.2
                if combined > best_combined:
                    best_item=item; best_cos=cos; best_combined=combined

            score = round(best_cos,6)
            query_results[q] = {
                "best_code": safe_str(best_item.get("ma_hieu")),
                "score":     score,
                "label":     "true" if score > SIMILARITY_AUDIT_THRESHOLD else "false",
                "best_item": best_item,
            }
        for row_idx, key in row_to_key.items():
            if row_idx not in results:
                results[row_idx] = query_results.get(key,{"best_code":"","score":0.0,"label":"false","best_item":None})

    return results

def find_exact_code_match(item_code: str, source_item_name: str="", source_desc: str="") -> Optional[dict]:
    knowledge      = load_knowledge_index()
    normalized_code= normalize_ma_hieu(item_code)
    if not normalized_code:
        normalized_code = extract_ma_hieu_from_text(item_code, source_item_name, source_desc)
    if not normalized_code: return None
    exact = knowledge["by_code"].get(normalized_code)
    return build_match_payload(exact, 1.0, "ma_hieu_exact") if exact else None

def find_exact_name_match(source_item_name: str) -> Optional[dict]:
    knowledge = load_knowledge_index()
    exact_name = knowledge.get("by_exact_name", {}).get(_exact_name_key(source_item_name))
    return build_match_payload(exact_name, 1.0, "ten_cong_viec_exact") if exact_name else None

def select_knowledge_match_for_ai_from_audit(
    sim_audit: dict, item_code: str,
    source_item_name: str="", source_desc: str="",
) -> Tuple[Optional[dict], Optional[dict], dict]:
    exact_match = find_exact_code_match(item_code, source_item_name, source_desc)
    exact_name_match = find_exact_name_match(source_item_name)
    reference_match = None
    if sim_audit.get("best_item") is not None:
        reference_match = build_match_payload(sim_audit["best_item"], sim_audit.get("score",0.0), "best_similarity_code")
    if exact_match: return exact_match, reference_match, sim_audit
    if exact_name_match: return exact_name_match, reference_match, sim_audit
    if reference_match and float(sim_audit.get("score",0.0)) > SIMILARITY_AUDIT_THRESHOLD:
        return reference_match, reference_match, sim_audit
    return None, reference_match, sim_audit


# =========================
# AI DESCRIPTION HELPERS
# =========================
def _append_brand_to_ai_desc(ai_text: str, source_desc: str, nhan_hieu: str = "") -> str:
    """Bổ sung nhãn hiệu/xuất xứ từ source_desc vào cuối ai_desc nếu chưa có."""
    if not source_desc and not nhan_hieu:
        return ai_text
    STOP_AHEAD = (
        r"(?=\s*(?:Nhãn hiệu|Thương hiệu|Hãng sản xuất|Hãng|NSX"
        r"|Xuất xứ|Mã hiệu|Đơn vị sản xuất|Đơn vị nhập khẩu)\s*[:/]|[.\n]|$)"
    )
    FLAGS = re.IGNORECASE | re.DOTALL
    # "Nhãn hiệu / Xuất xứ: X / Y" — combined field, capture everything after colon
    pat_gop = (
        r"(?:Nhãn hiệu|Thương hiệu)\s*/\s*[Xx]uất xứ\s*[:/]\s*(.+?)" + STOP_AHEAD
    )
    pat_nhan = (
        r"(?:Nhãn hiệu|Thương hiệu|Hãng sản xuất"
        r"|(?<!\w)Hãng|(?<!\w)NSX|Đơn vị sản xuất|Đơn vị nhập khẩu)"
        r"\s*[:/]\s*(.+?)" + STOP_AHEAD
    )
    pat_xuat = r"[Xx]uất xứ\s*[:/]\s*(.+?)" + STOP_AHEAD

    def _first(pat, text):
        m = re.search(pat, text, FLAGS)
        if not m:
            return None
        v = m.group(1).strip().rstrip(".,")
        return v or None

    if nhan_hieu and nhan_hieu.strip():
        nhan = nhan_hieu.strip()
        xuat = _first(pat_xuat, source_desc)
    else:
        gop = _first(pat_gop, source_desc)
        if gop:
            nhan = gop
            xuat = None
        else:
            nhan = _first(pat_nhan, source_desc)
            xuat = _first(pat_xuat, source_desc)

    if not nhan and not xuat:
        return ai_text

    ai_lower = ai_text.lower()
    parts = []
    if nhan and nhan.lower() not in ai_lower:
        parts.append(f"Nhãn hiệu: {nhan}")
    if xuat and xuat.lower() not in ai_lower:
        parts.append(f"Xuất xứ: {xuat}")
    if not parts:
        return ai_text

    text = ai_text.rstrip()
    if text and text[-1] not in ".!?":
        text += "."
    return f"{text} {'. '.join(parts)}."


def normalize_ai_description(text: str) -> str:
    text = normalize_spaces(text)
    if not text: return ""
    text = re.sub(r"^\s*[-•*]+\s*","",text)
    text = re.sub(r"^\s*\d+[.)]\s*","",text)
    text = text.strip().strip('"').strip("'")
    if text and text[-1] not in ".!?": text += "."
    return capitalize_first(text)

def guess_code_token_meaning(token: str, is_last: bool = False) -> Optional[str]:
    token = normalize_ma_hieu(token)
    if not token: return None
    if re.fullmatch(r"B(\d+)F",token): return f"tầng hầm {re.fullmatch(r'B(\d+)F',token).group(1)}"
    if token=="BF": return "tầng hầm"
    if token=="GF": return "tầng trệt"
    if token=="LG": return "tầng trệt thấp"
    if token=="UG": return "tầng trệt trên"
    if token=="RF": return "tầng mái"
    if token=="PH": return "tầng penthouse"
    if re.fullmatch(r"(\d+)F",token): return f"tầng {re.fullmatch(r'(\d+)F',token).group(1)}"
    if re.fullmatch(r"L(\d+)",token): return f"tầng {re.fullmatch(r'L(\d+)',token).group(1)}"
    if re.fullmatch(r"[A-Z]+\d+",token): return f"khối {token}"
    if re.fullmatch(r"T\d+",token):       return f"tháp {token}"
    if re.fullmatch(r"[A-Z]{2,8}",token):
        return f"khu vực {token}" if is_last else token
    return None

def choose_main_object_phrase(code: str, item_name_vi: str, primary_match: Optional[dict]) -> str:
    tokens = split_code_tokens(code)
    if tokens and tokens[0] in PANEL_PREFIX_MAP: return PANEL_PREFIX_MAP[tokens[0]]
    if primary_match:
        km_name = normalize_vi_text(fallback_translate_item_name(
            safe_str(primary_match.get("ten_cong_viec")), safe_str(primary_match.get("description"))))
        if km_name: return km_name
    item_name_vi = normalize_vi_text(item_name_vi)
    if item_name_vi: return item_name_vi
    if tokens and tokens[0]: return f"Hạng mục mã {tokens[0]}"
    return "Hạng mục MEP"

def build_compact_code_meaning(
    item_name_vi: str, source_item_name: str, source_desc: str,
    row_item_code: str, primary_match: Optional[dict] = None,
) -> str:
    row_code    = normalize_ma_hieu(row_item_code) or extract_ma_hieu_from_text(row_item_code, source_item_name, source_desc)
    primary_code= normalize_ma_hieu(primary_match.get("ma_hieu","")) if primary_match else ""
    final_code  = primary_code or row_code
    main_obj    = choose_main_object_phrase(final_code, item_name_vi, primary_match)
    if not final_code: return normalize_ai_description(f"{main_obj}, theo bản vẽ và yêu cầu kỹ thuật.")
    tokens = split_code_tokens(final_code)
    parts  = [p for i, t in enumerate(tokens[1:],1)
              if (p := guess_code_token_meaning(t, is_last=(i==len(tokens)-1)))]
    parts  = dedupe_preserve_order(parts)
    if parts: return normalize_ai_description(f"{main_obj}, tại {', '.join(parts)}.")
    return normalize_ai_description(f"{main_obj}, theo mã {final_code}.")

def build_local_mep_ai_description(
    item_name_vi: str, source_item_name: str="", source_desc: str="",
    row_item_code: str="", primary_match: Optional[dict]=None,  # noqa: ARG001
    translated_desc_vi: str="",
) -> str:
    """Fallback AI desc: tên hạng mục + mô tả đặc điểm + xuất xứ.
    Chỉ dùng primary_match khi score >= threshold (đã lọc trước khi truyền vào).
    """
    item_vi = normalize_vi_text(item_name_vi) or fallback_translate_item_name(source_item_name, source_desc)
    if not item_vi:
        item_vi = normalize_inline(source_item_name) or "Hạng mục MEP"

    # Ưu tiên dùng bản dịch tiếng Việt (translated_desc_vi) thay vì source_desc
    _desc_src = translated_desc_vi or source_desc
    desc_vi = normalize_spaces(_desc_src)
    if desc_vi and not still_has_foreign(desc_vi):
        desc_part = normalize_vi_text(desc_vi)
    elif translated_desc_vi and not still_has_foreign(translated_desc_vi):
        desc_part = normalize_vi_text(translated_desc_vi)
    elif desc_vi:
        desc_part = fallback_translate_desc(desc_vi, source_item_name)
    else:
        desc_part = ""

    # Nếu có primary_match (đã qua ngưỡng 88%) → dùng thông tin từ knowledge
    if primary_match:
        km_desc = normalize_vi_text(safe_str(primary_match.get("description", "")))
        if km_desc and (not desc_part or len(desc_part) < 20):
            desc_part = km_desc

    # Loại bỏ tên B trùng lặp bên trong desc_part
    if desc_part and item_vi:
        item_cmp = normalize_basic_compare(item_vi)
        # Tách desc thành các câu, bỏ câu nào trùng khớp với tên B
        sentences = [s.strip() for s in re.split(r'(?<=\.)\s+', desc_part) if s.strip()]
        sentences = [s for s in sentences if normalize_basic_compare(s.rstrip(".")) != item_cmp]
        desc_part = " ".join(sentences).strip()

    # Ghép: tên + mô tả (không lặp)
    parts = [item_vi]
    if desc_part and normalize_basic_compare(desc_part) != normalize_basic_compare(item_vi):
        parts.append(desc_part)

    result = ". ".join(p.rstrip(".") for p in parts if p) + "."
    return ensure_components_description_format(result, primary_match)


# =========================
# ASYNC CLAUDE API
# =========================
def can_use_claude() -> bool:
    return bool(ANTHROPIC_API_KEY)

def extract_response_text(data: dict) -> str:
    blocks = data.get("content",[]) or []
    return "".join(b.get("text","") for b in blocks if isinstance(b,dict) and b.get("type")=="text").strip()

def _make_aiohttp_headers() -> dict:
    return {
        "x-api-key":         ANTHROPIC_API_KEY,
        "anthropic-version": ANTHROPIC_VERSION,
        "content-type":      "application/json",
    }

async def _post_claude_async(
    session:      aiohttp.ClientSession,
    semaphore:    asyncio.Semaphore,
    system_prompt:str,
    user_prompt:  str,
    max_tokens:   int  = 4000,
    model:        str  = None,
    max_retries:  int  = None,
) -> str:
    _model   = model or CLAUDE_MODEL
    _retries = max_retries if max_retries is not None else MAX_RETRIES
    payload  = {
        "model": _model, "max_tokens": max_tokens, "temperature": 0.1,
        "system": system_prompt,
        "messages": [{"role":"user","content":user_prompt}],
    }
    last_error = None

    async with semaphore:
        for attempt in range(1, _retries + 1):
            try:
                timeout = aiohttp.ClientTimeout(total=REQUEST_TIMEOUT)
                async with session.post(CLAUDE_API_URL, json=payload, timeout=timeout) as resp:
                    if resp.status == 429:
                        raw_ra  = resp.headers.get("retry-after","")
                        try:   wait_s = float(raw_ra)
                        except Exception: wait_s = min(MAX_RETRY_WAIT, 5*(2**(attempt-1)))
                        import random; wait_s = min(MAX_RETRY_WAIT, wait_s + random.uniform(0,2))
                        print(f"[429] attempt {attempt}/{_retries} model={_model}, wait {wait_s:.1f}s...")
                        await asyncio.sleep(wait_s); continue
                    if resp.status in {500,502,503,504,529}:
                        wait_s = min(MAX_RETRY_WAIT, 3*attempt)
                        await asyncio.sleep(wait_s); continue
                    resp.raise_for_status()
                    data = await resp.json()
                    return extract_response_text(data)
            except asyncio.TimeoutError:
                last_error = f"Timeout {REQUEST_TIMEOUT}s"
                if attempt < _retries: await asyncio.sleep(min(MAX_RETRY_WAIT,10*attempt))
            except aiohttp.ClientError as exc:
                last_error = exc
                if attempt < _retries: await asyncio.sleep(min(MAX_RETRY_WAIT,5*attempt))
            except Exception as exc:
                last_error = exc
                if attempt >= _retries: break
                await asyncio.sleep(min(MAX_RETRY_WAIT,5*attempt))

    raise RuntimeError(f"Claude async thất bại {_retries} lần. {last_error}")


_last_api_call_time: float = 0.0

def _throttle_api_call() -> None:
    global _last_api_call_time
    elapsed = time.time() - _last_api_call_time
    if elapsed < MIN_DELAY_BETWEEN_CALLS:
        time.sleep(MIN_DELAY_BETWEEN_CALLS - elapsed)
    _last_api_call_time = time.time()

def post_claude_text(system_prompt: str, user_prompt: str,
                     max_tokens: int = 4000, model: str = None) -> str:
    if not can_use_claude(): raise RuntimeError("ANTHROPIC_API_KEY trống.")
    _model  = model or CLAUDE_MODEL
    payload = {
        "model": _model, "max_tokens": max_tokens, "temperature": 0.1,
        "system": system_prompt,
        "messages": [{"role":"user","content":user_prompt}],
    }
    last_error = None
    for attempt in range(1, MAX_RETRIES + 1):
        _throttle_api_call()
        try:
            response = SESSION.post(CLAUDE_API_URL, json=payload, timeout=REQUEST_TIMEOUT)
            if response.status_code == 429:
                import random
                raw_ra = response.headers.get("retry-after","")
                try:   wait_s = float(raw_ra)
                except Exception: wait_s = min(MAX_RETRY_WAIT, 5*(2**(attempt-1)))
                wait_s = min(MAX_RETRY_WAIT, wait_s + random.uniform(0,2))
                print(f"[429] sync attempt {attempt}, wait {wait_s:.1f}s..."); time.sleep(wait_s); continue
            if response.status_code in {500,502,503,504,529}:
                time.sleep(min(MAX_RETRY_WAIT,3*attempt)); continue
            response.raise_for_status()
            return extract_response_text(response.json())
        except requests.exceptions.Timeout:
            last_error = "Timeout"; time.sleep(min(MAX_RETRY_WAIT,10*attempt))
        except Exception as exc:
            last_error = exc
            if attempt >= MAX_RETRIES: break
            time.sleep(min(MAX_RETRY_WAIT,5*attempt))
    raise RuntimeError(f"Claude sync thất bại {MAX_RETRIES} lần. {last_error}")


# =========================
# TRANSLATION HELPERS
# =========================
def get_translation_cache_key(row: dict) -> Tuple[str,str,str,str]:
    return (
        normalize_inline(row.get("source_item_name","")),
        normalize_spaces(row.get("source_desc_cleaned","")),
        normalize_inline(row.get("item_name_vi_seed","")),
        normalize_spaces(row.get("translated_desc_vi_seed","")),
    )

def build_translation_batch_payload(batch_rows: List[dict],
                                    model_hint: str = "haiku") -> Tuple[str,str]:
    rows_payload = [{
        "row":                     row["row_idx"],
        "source_item_name":        row["source_item_name"],
        "source_desc":             row["source_desc_cleaned"],
    } for row in batch_rows]

    system_prompt = (
        "Bạn là kỹ sư MEP Việt Nam. Dịch TOÀN BỘ nội dung sang TIẾNG VIỆT.\n\n"
        "QUY TẮC BẮT BUỘC:\n"
        "1. item_name_vi: dịch source_item_name sang tiếng Việt 100%\n"
        "2. translated_desc_vi: dịch source_desc sang tiếng Việt 100%\n"
        "3. KHÔNG ĐƯỢC để nguyên tiếng Anh hay bất kỳ ngoại ngữ nào\n"
        "4. CHỈ giữ nguyên: mã sản phẩm, model, số liệu kỹ thuật (DN, kV, mm², kW, IP), nhãn hiệu, xuất xứ\n"
        "5. Không bịa thông số, không thêm bớt nội dung\n\n"
        "VÍ DỤ:\n"
        '- "LED batten surface-mounted acrylic diffuser" → "Đèn LED batten gắn nổi chụp tán xạ acrylic"\n'
        '- "Supply and install light fittings with all necessary fixing accessories" → '
        '"Cung cấp và lắp đặt bộ đèn chiếu sáng bao gồm đầy đủ phụ kiện lắp đặt cần thiết"\n'
        '- "Gate valve DN50 PN16" → "Van cổng DN50 PN16"\n\n'
        'Trả về JSON: {"results":[{"row":2,"item_name_vi":"...","translated_desc_vi":"..."}]}'
    )
    return system_prompt, json.dumps({"rows": rows_payload}, ensure_ascii=False)


def _apply_translation_results_to_batch(batch_rows: List[dict], llm_map: Dict[int, dict]) -> None:
    global _TRANSLATION_DIRTY
    for row in batch_rows:
        llm_result = llm_map.get(row["row_idx"], {})
        item_name_vi    = normalize_vi_text(llm_result.get("item_name_vi",""))
        translated_desc = normalize_vi_text(llm_result.get("translated_desc_vi",""))
        if row.get("need_desc_vi"):
            print(f"  [DEBUG] row {row['row_idx']}: llm_result={bool(llm_result)}, "
                  f"translated_desc='{translated_desc[:50]}'" if translated_desc else
                  f"  [DEBUG] row {row['row_idx']}: llm_result={bool(llm_result)}, translated_desc=RỖNG")

        if not item_name_vi or (row.get("need_item_name_vi") and not looks_like_vietnamese(item_name_vi)
                                and not is_technical_code(row["source_item_name"])):
            item_name_vi = fallback_translate_item_name(row["source_item_name"], row["source_desc_cleaned"])

        if not translated_desc or (row.get("need_desc_vi") and not looks_like_vietnamese(translated_desc)
                                    and not is_technical_code(row["source_desc_cleaned"])):
            translated_desc = fallback_translate_desc(row["source_desc_cleaned"], row["source_item_name"])

        cache_key = get_translation_cache_key(row)
        _name_ok  = not still_has_foreign(item_name_vi) and (item_name_vi or not row.get("need_item_name_vi"))
        _desc_ok  = not still_has_foreign(translated_desc) and (translated_desc or not row.get("need_desc_vi"))
        if _name_ok and _desc_ok:
            TRANSLATION_CACHE[cache_key] = (item_name_vi, translated_desc)
            _INST_TRANSLATION_KEYS.add(cache_key)
            _TRANSLATION_DIRTY = True
        else:
            print(f"  [SKIP CACHE] row {row['row_idx']}: kết quả dịch vẫn tiếng Anh")

        final_name_vi = item_name_vi if row.get("need_item_name_vi") else row["existing_item_name_vi"]
        final_desc_vi = translated_desc if row.get("need_desc_vi") else row["existing_desc_vi"]

        if not normalize_inline(final_name_vi):
            final_name_vi = fallback_translate_item_name(row["source_item_name"], row["source_desc_cleaned"])
        if row.get("need_desc_vi") and not looks_like_vietnamese(final_desc_vi) \
                and not looks_like_vietnamese(row.get("source_desc_cleaned","")):
            final_desc_vi = fallback_translate_desc(row["source_desc_cleaned"], row["source_item_name"])

        row["final_item_name_vi"] = final_name_vi
        row["final_desc_vi"]      = final_desc_vi


# =========================
# AI DESCRIPTION HELPERS
# =========================
def get_ai_desc_cache_key(row: dict) -> Tuple:
    primary_match = row.get("primary_match") or {}
    components    = primary_match.get("components") or []
    return (
        normalize_inline(row.get("source_item_name","")),
        normalize_spaces(row.get("source_desc_cleaned","")),
        normalize_inline(row.get("final_item_name_vi","")),
        normalize_inline(primary_match.get("ten_cong_viec","")),
        normalize_spaces(primary_match.get("description","")),
        "|".join(format_component_entry(c) for c in components),
    )

def build_ai_description_batch_payload(batch_rows: List[dict],
                                        model_hint: str = "haiku") -> Tuple[str, str]:
    rows_payload = []
    for row in batch_rows:
        pm = row.get("primary_match") or {}
        rm = row.get("reference_match") or {}
        sa = row.get("sim_audit") or {}
        rows_payload.append({
            "row":                      row["row_idx"],
            "row_item_code":            safe_str(row.get("source_item_code")),
            "translated_item_name":     safe_str(row.get("final_item_name_vi")),
            "translated_desc":          safe_str(row.get("final_desc_vi")),
            "primary_match":            {"match_type":safe_str(pm.get("match_type")),
                                         "score":pm.get("score"),"ma_hieu":safe_str(pm.get("ma_hieu")),
                                         "ten_cong_viec":safe_str(pm.get("ten_cong_viec")),
                                         "description":safe_str(pm.get("description")),
                                         "embed_text":safe_str(pm.get("embed_text")),
                                         "components":pm.get("components") or []} if pm else None,
            "reference_similarity_match":{"match_type":safe_str(rm.get("match_type")),
                                           "score":rm.get("score"),"ma_hieu":safe_str(rm.get("ma_hieu")),
                                           "ten_cong_viec":safe_str(rm.get("ten_cong_viec")),
                                           "description":safe_str(rm.get("description")),
                                           "embed_text":safe_str(rm.get("embed_text")),
                                           "components":rm.get("components") or []} if rm else None,
            "best_similarity_code":     safe_str(sa.get("best_code")),
            "best_similarity_score":    sa.get("score"),
        })

    if model_hint == "haiku":
        system_prompt = (
            "Kỹ sư MEP Việt Nam. Viết mô tả kỹ thuật TIẾNG VIỆT cho hạng mục BOQ.\n\n"
            "BẮT BUỘC:\n"
            "- OUTPUT PHẢI là tiếng Việt (không được viết tiếng Anh)\n"
            "- Mỗi ai_desc PHẢI dài tối thiểu 20 ký tự\n"
            "- Mở đầu bằng translated_item_name, thêm thông số từ translated_desc\n"
            "- Giữ nguyên mã/model/DN/kV/mm²/IP. Không bịa thông số.\n\n"
            "- Phần mô tả sẽ phải có tên hạng mục + thông tin chi tiết + components(nếu có)+ xuất sứ\n"
            "QUY TẮC COMPONENTS:\n"
            "Nếu primary_match.components KHÔNG rỗng → PHẢI thêm vào CUỐI:\n"
            "Gồm : \"component 1; component 2\"\n\n"
            'JSON thuần: {"results":[{"row":2,"ai_desc":"..."}]}'
        )
    else:
        system_prompt = (
            "Bạn là kỹ sư MEP senior. Viết mô tả kỹ thuật tiếng Việt cho hạng mục BOQ.\n\n"
            "Phần mô tả sẽ phải có tên hạng mục + thông tin chi tiết + components(nếu có)+ xuất sứ\n"
            "NGUYÊN TẮC BẮT BUỘC — COMPONENTS:\n"
            "Nếu primary_match.components có dữ liệu → BẮT BUỘC thêm vào cuối:\n"
            "Gồm : \"component 1; component 2\"\n\n"
            "Nguồn theo thứ tự ưu tiên:\n"
            "1. translated_desc (giữ nguyên mã/model/DN/kV/mm²/IP/hãng)\n"
            "2. translated_item_name (câu mở đầu)\n"
            "3. primary_match.description (bổ sung nếu cần)\n"
            "4. primary_match.components (BẮT BUỘC cuối)\n\n"
            "Không bịa thông số, không markdown.\n"
            'JSON thuần: {"results":[{"row":2,"ai_desc":"..."}]}'
        )
    return system_prompt, json.dumps({"rows": rows_payload}, ensure_ascii=False)


def _ai_desc_needs_retry(ai_text: str, primary_match: Optional[dict]) -> bool:
    text = normalize_spaces(ai_text)
    if len(text) < 20: return True
    if not looks_like_vietnamese(text): return True
    components = (primary_match or {}).get("components") or []
    if components and not re.search(r'gồm\s*:\s*"', text, flags=re.IGNORECASE): return True
    return False


# =========================
# SIMILARITY (CPU + GPU)
# =========================
def resolve_similarity_and_matches_for_batch(batch: List[dict]) -> None:
    rows_need_sim = []
    for row in batch:
        sim_key = get_sim_cache_key(row["source_item_name"], row["source_desc_cleaned"])
        row["_sim_cache_key"] = sim_key
        if sim_key in SIM_AUDIT_CACHE:
            cached = SIM_AUDIT_CACHE[sim_key]
            pm, rm, sa = select_knowledge_match_for_ai_from_audit(
                sim_audit=cached, item_code=row["source_item_code"],
                source_item_name=row["source_item_name"], source_desc=row["source_desc_cleaned"])
            row["sim_audit"]=sa; row["primary_match"]=pm; row["reference_match"]=rm
        else:
            rows_need_sim.append(row)

    if rows_need_sim:
        def _build_query(row):
            parts = [p for p in [
                normalize_inline(row.get("final_item_name_vi","")),
                normalize_inline(row.get("final_desc_vi","")),
                normalize_inline(row.get("source_item_name","")),
            ] if p]
            return " | ".join(parts)

        row_to_name  = {r["row_idx"]: _build_query(r) for r in rows_need_sim}
        sim_audit_map= find_best_similarity_code_by_name_batch(row_to_name)

        for row in rows_need_sim:
            sa = sim_audit_map.get(row["row_idx"],
                                   {"best_code":"","score":0.0,"label":"false","best_item":None})
            _sim_key = row["_sim_cache_key"]
            SIM_AUDIT_CACHE[_sim_key] = {
                "best_code": safe_str(sa.get("best_code")),
                "score":     float(sa.get("score",0.0) or 0.0),
                "label":     safe_str(sa.get("label")),
                "best_item": make_minimal_item(sa.get("best_item")),
            }
            _INST_SIM_AUDIT_KEYS.add(_sim_key)
            pm, rm, sa = select_knowledge_match_for_ai_from_audit(
                sim_audit=sa, item_code=row["source_item_code"],
                source_item_name=row["source_item_name"], source_desc=row["source_desc_cleaned"])
            row["sim_audit"]=sa; row["primary_match"]=pm; row["reference_match"]=rm

    cache_hits = len(batch) - len(rows_need_sim)
    if cache_hits: print(f"  [SIM CACHE] {cache_hits}/{len(batch)} hit, {len(rows_need_sim)} computed.")


def check_components_batch(item_names: List[str]) -> Dict[str,bool]:
    if not REWRITE_AI_IF_HAS_COMPONENTS: return {n:False for n in item_names}
    knowledge = load_knowledge_index()
    pool = knowledge["similarity_pool_items"]
    if not pool or knowledge["similarity_pool_matrix"].shape[0]==0: return {n:False for n in item_names}
    unique = list({normalize_inline(n) for n in item_names if normalize_inline(n)})
    if not unique: return {n:False for n in item_names}
    qvecs        = embed_texts(unique, persist_query_cache=True)
    score_matrix = _cosine_sim_matrix(qvecs, knowledge)   # GPU-accelerated
    best_idxs    = np.argmax(score_matrix, axis=1)
    best_scores  = score_matrix[np.arange(score_matrix.shape[0]), best_idxs]
    query_result: Dict[str,bool] = {}
    for i, q in enumerate(unique):
        sc = float(best_scores[i])
        query_result[q] = len(pool[int(best_idxs[i])].get("components") or []) > 0 if sc > SIMILARITY_AUDIT_THRESHOLD else False
    return {n: query_result.get(normalize_inline(n),False) for n in item_names}


# =========================
# 3-PHASE ASYNC PIPELINE
# =========================
async def _phase1_translate_all_async(
    session:   aiohttp.ClientSession,
    semaphore: asyncio.Semaphore,
    batch_data: List[dict],
) -> None:
    global _TRANSLATION_DIRTY

    async def _translate_one(bd: dict) -> None:
        fresh = bd["fresh"]
        if not fresh: return

        rows_need = [r for r in fresh if r.get("need_item_name_vi") or r.get("need_desc_vi")]
        rows_no_translate = [r for r in fresh if r not in rows_need]
        for row in rows_no_translate:
            if "final_item_name_vi" in row: continue
            row["final_item_name_vi"] = row["existing_item_name_vi"] or \
                fallback_translate_item_name(row["source_item_name"], row["source_desc_cleaned"])
            row["final_desc_vi"]      = row["existing_desc_vi"] or \
                fallback_translate_desc(row["source_desc_cleaned"], row["source_item_name"])

        if not rows_need: return

        missing_rows: List[dict] = []
        for row in rows_need:
            cache_key = get_translation_cache_key(row)
            if cache_key in TRANSLATION_CACHE:
                cached_pair = TRANSLATION_CACHE[cache_key]
                nv, dv = cached_pair
                # Cache invalid nếu: kết quả còn ngoại ngữ, hoặc cần dịch mà rỗng
                _nv_bad = still_has_foreign(nv) or (row.get("need_item_name_vi") and not nv)
                _dv_bad = still_has_foreign(dv) or (row.get("need_desc_vi") and not dv)
                if _nv_bad or _dv_bad:
                    print(f"  [CACHE INVALID] row {row['row_idx']}: nv_bad={_nv_bad} dv_bad={_dv_bad} dv='{dv[:30] if dv else 'RỖNG'}'")
                    del TRANSLATION_CACHE[cache_key]; missing_rows.append(row)
                else:
                    print(f"  [CACHE HIT] row {row['row_idx']}: dv='{dv[:40] if dv else 'RỖNG'}'")
                    row["final_item_name_vi"] = nv if row.get("need_item_name_vi") else row["existing_item_name_vi"]
                    row["final_desc_vi"]      = dv if row.get("need_desc_vi")      else row["existing_desc_vi"]
            else:
                missing_rows.append(row)

        # Chia missing_rows thành sub-batches nhỏ (30 rows/call) chạy SONG SONG
        TRANSLATE_SUB_BATCH = 30
        llm_map: Dict[int, dict] = {}

        async def _translate_sub_batch(sb: List[dict]) -> None:
            sys_p, usr_p = build_translation_batch_payload(sb, model_hint="sonnet")
            try:
                raw = await _post_claude_async(session, semaphore, sys_p, usr_p,
                                               max_tokens=16000, model=CLAUDE_MODEL_TRANSLATE)
                parsed = safe_json_loads(raw)
                for item in parsed.get("results", []):
                    rn = item.get("row")
                    if rn is not None:
                        llm_map[int(rn)] = {
                            "item_name_vi":       normalize_vi_text(item.get("item_name_vi", "")),
                            "translated_desc_vi": normalize_vi_text(item.get("translated_desc_vi", "")),
                        }
            except Exception as exc:
                print(f"  [Translation sub-batch error] {exc}")

        if missing_rows:
            sub_batches = [missing_rows[i:i+TRANSLATE_SUB_BATCH]
                           for i in range(0, len(missing_rows), TRANSLATE_SUB_BATCH)]
            print(f"    → {len(missing_rows)} rows chia {len(sub_batches)} sub-batches song song")
            await asyncio.gather(*[_translate_sub_batch(sb) for sb in sub_batches])

        _apply_translation_results_to_batch(missing_rows, llm_map)

    total_rows_p1 = sum(
        1 for bd in batch_data
        for r in bd["fresh"]
        if r.get("need_item_name_vi") or r.get("need_desc_vi")
    )
    if total_rows_p1 == 0:
        print("  [SKIP] Phase 1: tất cả rows đã tiếng Việt, không cần dịch.")
        pbar1 = None
    else:
        pbar1 = _tqdm(total=total_rows_p1, desc="Phase 1 Dịch", unit="rows")

    async def _translate_one_tracked(bd: dict) -> None:
        n_need = sum(1 for r in bd["fresh"] if r.get("need_item_name_vi") or r.get("need_desc_vi"))
        await _translate_one(bd)
        if pbar1 and n_need: pbar1.update(n_need)

    await asyncio.gather(*[_translate_one_tracked(bd) for bd in batch_data])
    if pbar1: pbar1.close()
    if total_rows_p1 > 0:
        print(f"  ✓ Phase 1 xong: {total_rows_p1} rows dịch (Haiku)")


async def _ai_one(
    session:   aiohttp.ClientSession,
    semaphore: asyncio.Semaphore,
    bd: dict,
) -> None:
    global _AI_CACHE_DIRTY
    fresh = bd["fresh"]
    if not fresh: return

    rows_need_ai = [r for r in fresh if r.get("need_ai")]
    for row in fresh:
        if not row.get("need_ai"):
            row["final_ai"] = ensure_components_description_format(row["existing_ai"], row.get("primary_match"))

    if not rows_need_ai: return

    # Tách: đã match knowledge → fallback local | chưa match → gọi Sonnet
    rows_matched = []
    rows_api = []
    for row in rows_need_ai:
        ai_key = get_ai_desc_cache_key(row)
        row["ai_cache_key"] = ai_key
        if ai_key in AI_DESC_CACHE:
            row["final_ai"] = ensure_components_description_format(AI_DESC_CACHE[ai_key], row.get("primary_match"))
        elif row.get("primary_match"):
            # Đã match với knowledge → fallback local (tên + mô tả + Gồm nếu có components)
            rows_matched.append(row)
        else:
            # Hoàn toàn mới, không match gì → gọi Sonnet
            rows_api.append(row)

    for row in rows_matched:
        ai_text = build_local_mep_ai_description(
            item_name_vi=row.get("final_item_name_vi", ""),
            source_item_name=row.get("source_item_name", ""),
            source_desc=row.get("source_desc_cleaned", ""),
            row_item_code=row.get("source_item_code", ""),
            primary_match=row.get("primary_match"),
            translated_desc_vi=row.get("final_desc_vi", ""),
        )
        ai_key = row["ai_cache_key"]
        AI_DESC_CACHE[ai_key] = ai_text
        _INST_AI_DESC_KEYS.add(ai_key)
        _AI_CACHE_DIRTY = True
        row["final_ai"] = ai_text
    if rows_matched:
        print(f"  [LOCAL] {len(rows_matched)} rows đã match knowledge → fallback local", flush=True)

    # Chưa match → gọi Sonnet
    llm_map: Dict[int, str] = {}
    if rows_api:
        print(f"  [AI desc] Gọi Sonnet cho {len(rows_api)} rows mới...", flush=True)
        sys_p, usr_p = build_ai_description_batch_payload(rows_api, model_hint="sonnet")
        try:
            raw    = await _post_claude_async(session, semaphore, sys_p, usr_p,
                                               max_tokens=8192, model=CLAUDE_MODEL_AI_FALLBACK)
            parsed = safe_json_loads(raw)
            for item in parsed.get("results",[]):
                rn = item.get("row")
                if rn is not None:
                    llm_map[int(rn)] = item.get("ai_desc", "")
        except Exception as exc:
            print(f"  [AI desc Sonnet error] {exc} → fallback local", flush=True)

    for row in rows_api:
        ai_text = llm_map.get(row["row_idx"], "")
        if not ai_text or _ai_desc_needs_retry(ai_text, row.get("primary_match")):
            ai_text = build_local_mep_ai_description(
                item_name_vi=row.get("final_item_name_vi", ""),
                source_item_name=row.get("source_item_name", ""),
                source_desc=row.get("source_desc_cleaned", ""),
                row_item_code=row.get("source_item_code", ""),
                primary_match=row.get("primary_match"),
                translated_desc_vi=row.get("final_desc_vi", ""),
            )
        ai_key = row["ai_cache_key"]
        AI_DESC_CACHE[ai_key] = ai_text
        _INST_AI_DESC_KEYS.add(ai_key)
        _AI_CACHE_DIRTY = True
        row["final_ai"] = ai_text


async def process_all_batches_async(all_batches: List[List[dict]]) -> List[dict]:
    batch_data = []
    for batch in all_batches:
        cached = [r for r in batch if r.get("_from_output_cache")]
        fresh  = [r for r in batch if not r.get("_from_output_cache")]
        batch_data.append({"batch": batch, "cached": cached, "fresh": fresh})
    boq_component_lookup = load_boq_component_lookup()

    semaphore = asyncio.Semaphore(MAX_CONCURRENT_API_CALLS)
    headers   = _make_aiohttp_headers()

    async with aiohttp.ClientSession(headers=headers) as session:
        print(f"\n[Phase 1] Dịch {len(batch_data)} batches song song → Haiku ...")
        await _phase1_translate_all_async(session, semaphore, batch_data)

        print(f"[Phase 2] Similarity {len(batch_data)} batches (GPU={DEVICE}) ...")
        total_rows_p2 = sum(len(bd["fresh"]) for bd in batch_data)
        with _tqdm(total=total_rows_p2, desc="Phase 2 Similarity", unit="rows") as pbar2:
            for bd in batch_data:
                fresh = bd["fresh"]
                if not fresh:
                    continue
                resolve_similarity_and_matches_for_batch(fresh)
                for r in fresh:
                    r.setdefault("final_item_name_vi", r.get("existing_item_name_vi",""))
                    r.setdefault("final_desc_vi",      r.get("existing_desc_vi",""))
                    r.setdefault("final_ai",           r.get("existing_ai",""))
                    r.setdefault("sim_audit",          {"best_code":"","score":0.0,"label":"false","best_item":None})
                    r.setdefault("primary_match",  None)
                    r.setdefault("reference_match",None)
                pbar2.update(len(fresh))

        forced_component_recall = 0
        skipped_component_cached = 0
        for bd in batch_data:
            for r in bd["fresh"]:
                refreshed_components = refresh_row_components_from_boq(r, boq_component_lookup)
                has_components = bool(refreshed_components or (r.get("primary_match") or {}).get("components"))
                if not (REWRITE_AI_IF_HAS_COMPONENTS and has_components):
                    r["_force_component_api_retry"] = False
                    continue
                # Nếu AI desc đã có "Gồm :" với đúng components → skip, không cần gọi lại API
                existing_ai = normalize_spaces(r.get("final_ai") or r.get("existing_ai") or "")
                expected_suffix = build_components_suffix(
                    (r.get("primary_match") or {}).get("components") or [])
                if existing_ai and expected_suffix and expected_suffix in existing_ai:
                    r["_force_component_api_retry"] = False
                    skipped_component_cached += 1
                    continue
                r["_force_component_api_retry"] = True
                r["need_ai"] = True
                r["_ai_from_learned"] = False
                r["final_ai"] = ""
                forced_component_recall += 1
        if skipped_component_cached:
            print(f"  [COMPONENT CACHE] {skipped_component_cached} rows đã có components đúng → skip API.")
        if forced_component_recall:
            print(f"  [FORCE COMPONENT RECALL] {forced_component_recall} rows có components sẽ gọi lại API.")

        _learned_hits = 0
        for bd in batch_data:
            for r in bd["fresh"]:
                if r.get("_force_component_api_retry"):
                    continue
                if not r.get("need_ai") or r.get("final_ai"):
                    continue
                _lai = lookup_learned_ai(
                    r.get("source_item_name",""),
                    r.get("source_item_code",""),
                    r.get("final_item_name_vi",""),
                    r.get("source_desc_cleaned",""),
                )
                if _lai:
                    r["final_ai"]        = _lai
                    r["_ai_from_learned"]= True
                    _learned_hits       += 1
        if _learned_hits:
            print(f"  [LEARNED LOOKUP] {_learned_hits} rows lấy AI desc từ learned items.")

        rows_need_api = sum(
            1 for bd in batch_data
            for r in bd["fresh"]
            if r.get("need_ai")
            and not r.get("_ai_from_learned")
            and (r.get("_force_component_api_retry")
                 or (not r.get("final_ai") and get_ai_desc_cache_key(r) not in AI_DESC_CACHE))
        )
        if rows_need_api == 0:
            print("  [SKIP] Phase 3: tất cả AI desc đã cache.")
            for bd in batch_data:
                for r in bd["fresh"]:
                    if r.get("need_ai") and not r.get("final_ai"):
                        if r.get("_force_component_api_retry"):
                            r["final_ai"] = build_local_mep_ai_description(
                                item_name_vi=r.get("final_item_name_vi", ""),
                                source_item_name=r.get("source_item_name", ""),
                                source_desc=r.get("source_desc_cleaned", ""),
                                row_item_code=r.get("source_item_code", ""),
                                primary_match=r.get("primary_match"),
                                translated_desc_vi=r.get("final_desc_vi", ""),
                            )
                        else:
                            r["final_ai"] = AI_DESC_CACHE.get(get_ai_desc_cache_key(r), r.get("existing_ai",""))
                    else:
                        r.setdefault("final_ai", r.get("existing_ai",""))
        else:
            print(f"[Phase 3] {rows_need_api} rows cần API → Haiku+Sonnet ...")
            pbar3 = _tqdm(total=rows_need_api, desc="Phase 3 AI desc", unit="rows")

            async def _ai_one_tracked(bd: dict) -> None:
                n_before = sum(1 for r in bd["fresh"] if r.get("need_ai") and not r.get("final_ai"))
                try:
                    await _ai_one(session, semaphore, bd)
                except Exception as exc:
                    print(f"  [Phase 3 batch error] {exc} → fallback local")
                    for r in bd["fresh"]:
                        if r.get("need_ai") and not r.get("final_ai"):
                            r["final_ai"] = build_local_mep_ai_description(
                                item_name_vi=r.get("final_item_name_vi", ""),
                                source_item_name=r.get("source_item_name", ""),
                                source_desc=r.get("source_desc_cleaned", ""),
                                row_item_code=r.get("source_item_code", ""),
                                primary_match=r.get("primary_match"),
                                translated_desc_vi=r.get("final_desc_vi", ""),
                            )
                pbar3.update(n_before)

            await asyncio.gather(*[_ai_one_tracked(bd) for bd in batch_data])
            pbar3.close()
        print(f"  ✓ Phase 3 xong")

    all_rows: List[dict] = []
    for bd in batch_data:
        all_rows.extend(bd["batch"])
    return all_rows


# =========================
# AUTO LEARN
# =========================
def build_learned_ai_lookup() -> None:
    global LEARNED_AI_LOOKUP
    LEARNED_AI_LOOKUP = {}
    _fallback_re = re.compile(r"theo bản vẽ và yêu cầu kỹ thuật\.$|theo mã [A-Z0-9-]+\.$")

    for item in LEARNED_ITEMS_RAM:
        final_ai = normalize_spaces(item.get("meta", {}).get("final_ai", ""))
        if not final_ai or len(final_ai) < 20 or _fallback_re.search(final_ai):
            final_ai = normalize_spaces(item.get("description", ""))
            if not final_ai or len(final_ai) < 20:
                continue

        code = normalize_ma_hieu(item.get("ma_hieu", ""))
        if code:
            LEARNED_AI_LOOKUP[f"code::{code}"] = final_ai

        name_vi = _normalize_name_for_key(item.get("ten_cong_viec", ""))
        if name_vi:
            LEARNED_AI_LOOKUP[f"name::{sha1_text(name_vi[:80])}"] = final_ai

        embed_text = safe_str(item.get("embed_text", ""))
        if embed_text:
            parts = embed_text.split(" | ")
            src_name = _normalize_name_for_key(parts[0].strip())
            if src_name and src_name != name_vi:
                LEARNED_AI_LOOKUP[f"name::{sha1_text(src_name[:80])}"] = final_ai

        meta = item.get("meta") or {}
        _src_name = safe_str(meta.get("source_item_name", ""))
        _src_desc = normalize_spaces(safe_str(meta.get("source_desc_cleaned", "")))
        if not _src_name and embed_text:
            _src_name = embed_text.split(" | ")[0].strip()
        if _src_name:
            _oc_key = sha1_text(
                f"{normalize_inline(_src_name)}||{_src_desc}"[:300]
            )
            LEARNED_AI_LOOKUP[f"src::{_oc_key}"] = final_ai

    print(f"  LEARNED_AI_LOOKUP: {len(LEARNED_AI_LOOKUP)} keys")


def lookup_learned_ai(source_item_name: str, source_item_code: str = "",
                      final_item_name_vi: str = "",
                      source_desc_cleaned: str = "") -> str:
    code = normalize_ma_hieu(source_item_code)
    if code:
        ai = LEARNED_AI_LOOKUP.get(f"code::{code}", "")
        if ai: return ai

    if source_item_name:
        _oc_key = sha1_text(
            f"{normalize_inline(source_item_name)}||{normalize_spaces(source_desc_cleaned)}"[:300]
        )
        ai = LEARNED_AI_LOOKUP.get(f"src::{_oc_key}", "")
        if ai: return ai

    for name in [final_item_name_vi, source_item_name]:
        if not name: continue
        key = f"name::{sha1_text(_normalize_name_for_key(name)[:80])}"
        ai = LEARNED_AI_LOOKUP.get(key, "")
        if ai: return ai

    return ""


def build_embed_text_for_learned_item(
    source_item_code: str, final_item_name_vi: str, final_desc_vi: str,
    final_ai: str, source_item_name: str, source_desc_cleaned: str,
) -> str:
    parts = []
    name  = _normalize_name_for_key(final_item_name_vi or source_item_name)
    if name: parts.append(name)
    desc  = normalize_spaces(final_desc_vi) or normalize_spaces(source_desc_cleaned)
    if desc: parts.append(desc)
    code  = normalize_ma_hieu(source_item_code)
    if code: parts.append(code)
    return " | ".join([p for p in parts if p])

def should_auto_learn_row(row: dict) -> bool:
    if not AUTO_LEARN_ENABLED: return False
    if row.get("_from_output_cache"): return False

    row_idx           = row.get("row_idx","?")
    source_item_code  = normalize_ma_hieu(row.get("source_item_code",""))

    if source_item_code:
        if f"code::{source_item_code}" in LEARNED_KEYS_CACHE: return False
    for _name_src in [row.get("final_item_name_vi",""), row.get("source_item_name","")]:
        if _name_src and f"name::{sha1_text(_normalize_name_for_key(_name_src)[:80])}" in LEARNED_KEYS_CACHE:
            return False

    if not source_item_code and not AUTO_LEARN_CODELESS_ROWS: return False

    final_item_name_vi = normalize_inline(row.get("final_item_name_vi",""))
    final_desc_vi      = normalize_spaces(row.get("final_desc_vi",""))
    final_ai           = normalize_spaces(row.get("final_ai",""))

    if not final_item_name_vi and not final_desc_vi and not final_ai: return False
    if final_ai and len(final_ai.strip()) < 20: final_ai = ""
    if final_ai and re.search(r"theo bản vẽ và yêu cầu kỹ thuật\.$|theo mã [A-Z0-9-]+\.$", final_ai.strip()): final_ai = ""
    if not final_item_name_vi and not final_desc_vi and not final_ai: return False
    if not final_item_name_vi: return False

    if row.get("need_item_name_vi") and looks_like_vietnamese(final_item_name_vi):
        return True

    primary_match = row.get("primary_match") or {}
    sim_audit     = row.get("sim_audit") or {}
    score         = float(sim_audit.get("score",0.0) or 0.0)
    match_type    = safe_str(primary_match.get("match_type"))

    if match_type == "ma_hieu_exact": return True
    if score > AUTO_LEARN_MIN_SCORE:  return True
    return False

def build_learned_item_from_row(row: dict) -> Optional[dict]:
    if not should_auto_learn_row(row): return None
    source_item_code   = normalize_ma_hieu(row.get("source_item_code",""))
    source_item_name   = safe_str(row.get("source_item_name",""))
    source_desc_cleaned= safe_str(row.get("source_desc_cleaned",""))
    final_item_name_vi = safe_str(row.get("final_item_name_vi",""))
    final_desc_vi      = safe_str(row.get("final_desc_vi",""))
    final_ai           = safe_str(row.get("final_ai",""))
    if re.search(r"theo bản vẽ và yêu cầu kỹ thuật\.$|theo mã [A-Z0-9-]+\.$", final_ai.strip()):
        final_ai = ""
    sim_audit    = row.get("sim_audit") or {}
    primary_match= row.get("primary_match") or {}
    learned_from = "row_result"
    if safe_str(primary_match.get("match_type"))=="ma_hieu_exact":   learned_from="exact_reference"
    elif float(sim_audit.get("score",0.0) or 0.0) > AUTO_LEARN_MIN_SCORE: learned_from="high_similarity"
    elif row.get("need_item_name_vi") and looks_like_vietnamese(final_item_name_vi): learned_from="translated_foreign"

    item_name   = normalize_vi_text(final_item_name_vi) or fallback_translate_item_name(source_item_name, source_desc_cleaned)
    description = normalize_vi_text(final_desc_vi)      or fallback_translate_desc(source_desc_cleaned, source_item_name)
    embed_text  = build_embed_text_for_learned_item(
        source_item_code, item_name, description, final_ai, source_item_name, source_desc_cleaned)
    if not item_name and not description and not embed_text: return None

    matched_components = primary_match.get("components") or [] if primary_match else []
    return {
        "id": f"learned::{sha1_text('||'.join([source_item_code,item_name,description,embed_text]))}",
        "ma_hieu": source_item_code, "ten_cong_viec": item_name,
        "description": description, "embed_text": embed_text,
        "components": matched_components, "source": "learned",
        "learned_from": learned_from, "learned_at": now_utc_iso(),
        "instance_id": INSTANCE_ID,
        "meta": {
            "best_similarity_score": float(sim_audit.get("score",0.0) or 0.0),
            "row_idx": int(row.get("row_idx",0) or 0),
            "final_ai": final_ai,
            "source_item_name":    source_item_name,
            "source_desc_cleaned": source_desc_cleaned,
        },
    }

def auto_learn_from_batch(batch: List[dict]) -> int:
    if not AUTO_LEARN_ENABLED: return 0
    global LEARNED_ITEMS_RAM, _LEARNED_ITEMS_DIRTY
    by_key: Dict[str,dict] = {build_learned_item_key(it): it for it in LEARNED_ITEMS_RAM}
    added = updated = 0
    for row in batch:
        if row.get("_from_output_cache"):
            continue
        item = build_learned_item_from_row(row)
        if not item: continue
        key = build_learned_item_key(item)
        if key in LEARNED_KEYS_CACHE: continue
        old = by_key.get(key)
        if old is None:
            by_key[key]=item; added+=1
        else:
            def _is_fallback(t): return bool(re.search(r"theo bản vẽ|theo mã [A-Z0-9-]+\.$", normalize_spaces(t)))
            new_ai = normalize_spaces(item.get("meta",{}).get("final_ai",""))
            old_ai = normalize_spaces(old.get("meta",{}).get("final_ai",""))
            new_from_claude = bool(new_ai) and not _is_fallback(new_ai)
            old_from_claude = bool(old_ai) and not _is_fallback(old_ai)
            new_desc = normalize_spaces(item.get("description",""))
            old_desc = normalize_spaces(old.get("description",""))
            if ((new_desc != old_desc) and
                ((new_from_claude and not old_from_claude) or
                 (new_from_claude==old_from_claude and len(new_desc)>len(old_desc)))):
                merged = dict(old); merged.update(item); by_key[key]=merged; updated+=1

    if added+updated > 0:
        LEARNED_ITEMS_RAM   = list(by_key.values())
        _LEARNED_ITEMS_DIRTY= True
        LEARNED_KEYS_CACHE.update(by_key.keys())
        print(f"Auto-learn: +{added} mới, ~{updated} cập nhật (RAM: {len(LEARNED_ITEMS_RAM)})")
    return added+updated

def maybe_refresh_knowledge_index_after_learning(n: int) -> None:
    global KNOWLEDGE_INDEX, _GPU_SIM_MATRIX, _GPU_SIM_MATRIX_SIG
    if n<=0 or not INCLUDE_LEARNED_IN_CURRENT_RUN: return
    print("Refresh knowledge index + GPU matrix sau học mới...")
    KNOWLEDGE_INDEX = None
    _GPU_SIM_MATRIX = None
    _GPU_SIM_MATRIX_SIG = ""
    load_knowledge_index()

def purge_ai_cache_entries_with_components() -> int:
    if not REWRITE_AI_IF_HAS_COMPONENTS: return 0
    knowledge    = load_knowledge_index()
    current_sig  = knowledge["signature"]
    try:
        last_sig = PURGE_SIGNATURE_FILE.read_text(encoding="utf-8").strip()
        if last_sig == current_sig:
            print("Purge skip: knowledge signature không đổi."); return 0
    except FileNotFoundError: pass

    pool = knowledge["similarity_pool_items"]
    items_with_comp = [it for it in pool if it.get("components")]
    if not items_with_comp: return 0
    comp_texts = set()
    for it in items_with_comp:
        for field in ["ten_cong_viec","embed_text"]:
            t = normalize_inline(safe_str(it.get(field,"")))
            if t: comp_texts.add(t)

    to_delete = []
    for key in AI_DESC_CACHE:
        if not isinstance(key, tuple): continue
        components_str = key[6] if len(key)>=7 else (key[5] if len(key)>=6 else "")
        match_ten      = normalize_inline(key[4]) if len(key)>=5 else ""
        if components_str or match_ten in comp_texts: to_delete.append(key)

    for key in to_delete:
        del AI_DESC_CACHE[key]
        _INST_AI_DESC_KEYS.discard(key)
    if to_delete:
        print(f"Purge {len(to_delete)} AI cache entries liên quan đến components.")
        write_json_file_atomic(AI_DESC_CACHE_FILE, [
            {"key":list(k),"value":v} for k,v in AI_DESC_CACHE.items()
            if k in _INST_AI_DESC_KEYS
        ])
    ensure_parent_dir(PURGE_SIGNATURE_FILE)
    PURGE_SIGNATURE_FILE.write_text(current_sig, encoding="utf-8")
    return len(to_delete)


# =========================
# SAFE SAVE + CHECKPOINT
# =========================
def _safe_wb_save(wb, output_path: str, label: str="") -> str:
    try:
        wb.save(output_path); print(f"Đã lưu {label}: {output_path}"); return output_path
    except PermissionError:
        backup = str(Path(output_path).parent / f"{Path(output_path).stem}_bak{int(time.time())}{Path(output_path).suffix}")
        wb.save(backup); print(f"[WARNING] Bị lock → lưu sang: {backup}"); return backup

def load_checkpoint(input_file: str) -> int:
    data = read_json_file(CHECKPOINT_FILE, {})
    return int(data.get("last_row",0)) if data.get("input_file")==input_file else 0

def save_checkpoint(input_file: str, last_row: int) -> None:
    write_json_file_atomic(CHECKPOINT_FILE, {"input_file":input_file,"last_row":last_row,"saved_at":now_utc_iso()})

def clear_checkpoint() -> None:
    if CHECKPOINT_FILE.exists(): CHECKPOINT_FILE.unlink()

def _retry_translate_single(text: str, context_type: str = "item_name") -> str:
    if not can_use_claude() or not text: return ""
    try:
        if context_type == "item_name":
            sys_p = "Bạn là kỹ sư MEP. Dịch tên hạng mục sau sang tiếng Việt kỹ thuật MEP. Chỉ trả về bản dịch."
        else:
            sys_p = "Bạn là kỹ sư MEP. Dịch mô tả kỹ thuật sau sang tiếng Việt. Giữ nguyên mã, thông số. Chỉ trả về bản dịch."
        return normalize_vi_text(post_claude_text(sys_p, text, max_tokens=2000,
                                                   model=CLAUDE_MODEL_TRANSLATE).strip())
    except Exception as exc:
        print(f"  [RETRY TRANSLATE] {exc}"); return ""


# =========================
# MAIN
# =========================
def main():
    if not os.path.exists(INPUT_FILE):      raise FileNotFoundError(f"Không tìm thấy input: {INPUT_FILE}")
    if not os.path.exists(KNOWLEDGE_JSON_FILE): raise FileNotFoundError(f"Không tìm thấy knowledge: {KNOWLEDGE_JSON_FILE}")
    if not ANTHROPIC_API_KEY:
        print("[LỖI] ANTHROPIC_API_KEY chưa set!"); raise SystemExit(1)

    print(f"Instance:      {INSTANCE_ID}")
    print(f"Device:        {DEVICE}  ({GPU_NAME})")
    print(f"Embed batch:   {EMBED_BATCH_SIZE}  (tự động theo GPU)")
    print(f"Row batch:     {BATCH_SIZE}  (tự động theo GPU)")
    print(f"Model dịch:    {CLAUDE_MODEL_TRANSLATE}")
    print(f"Model AI desc: {CLAUDE_MODEL_AI_DESC} (fallback: {CLAUDE_MODEL_AI_FALLBACK})")
    print(f"Concurrent:    {MAX_CONCURRENT_API_CALLS}")

    load_persistent_state()
    build_learned_ai_lookup()
    load_knowledge_index()
    purge_ai_cache_entries_with_components()

    is_json_input = Path(INPUT_FILE).suffix.lower() == ".json"
    _json_data: list = []
    _json_idx_map: Dict[int, dict] = {}
    col_map: Dict[str, Any] = {}
    _wb_write = None; _ws_write = None

    if is_json_input:
        raw_json = read_json_file(Path(INPUT_FILE), {})
        _json_data = raw_json.get("data", raw_json) if isinstance(raw_json, dict) else raw_json
        all_pms = {item["phan_muc"] for item in _json_data if item.get("phan_muc")}
        leaf_pms = {pm for pm in all_pms if not any(o.startswith(pm + ".") for o in all_pms if o != pm)}
        raw_rows = []
        for idx, item in enumerate(_json_data):
            pm = item.get("phan_muc", "")
            if pm and pm not in leaf_pms:
                continue
            _json_idx_map[idx] = item
            raw_rows.append({
                "row_idx": idx, "source_item_code": pm,
                "source_item_name": item.get("item_name", ""),
                "source_desc": item.get("source_desc", ""),
                "existing_item_name_vi": item.get("translated_item_name", ""),
                "existing_desc_vi": item.get("translated_desc", ""),
                "existing_ai": item.get("ai_desc", ""),
                "source_nhan_hieu": item.get("nhan_hieu", ""),
            })
        print(f"Tổng dòng JSON: {len(_json_data)}, hạng mục lá cần xử lý: {len(raw_rows)}")
    else:
        readable_input = get_readable_workbook_path(INPUT_FILE)
        keep_vba       = Path(readable_input).suffix.lower() == ".xlsm"

        wb_ro  = load_workbook(readable_input, read_only=True, data_only=True)
        ws_ro  = wb_ro[SHEET_NAME] if (SHEET_NAME and SHEET_NAME in wb_ro.sheetnames) else wb_ro[wb_ro.sheetnames[0]]
        print(f"Sheet: {ws_ro.title}")
        col_map = get_column_map(ws_ro, read_only=True)
        print("Map cột:", col_map)

        _col_ic = col_map["item_code"];  _col_in = col_map["item_name"]
        _col_sd = col_map["source_desc"]; _col_dv = col_map["translated_item_name"]
        _col_ev = col_map["translated_desc"]; _col_ai = col_map["ai_desc"]
        _col_nh = col_map.get("nhan_hieu")
        _max_col= max(c for c in [_col_ic,_col_in,_col_sd,_col_dv,_col_ev,_col_ai] + ([_col_nh] if _col_nh else []) if c)

        def _cell_val(row_cells, col_idx):
            if col_idx is None or col_idx<1 or col_idx>len(row_cells): return ""
            cell = row_cells[col_idx-1]
            return "" if cell.value is None else safe_str(cell.value)

        raw_rows = []
        for row_idx, row_cells in enumerate(
            ws_ro.iter_rows(min_row=START_ROW, max_col=_max_col, values_only=False), start=START_ROW
        ):
            ic = _cell_val(row_cells,_col_ic); in_ = _cell_val(row_cells,_col_in)
            sd = _cell_val(row_cells,_col_sd); dv  = _cell_val(row_cells,_col_dv)
            ev = _cell_val(row_cells,_col_ev); ai  = _cell_val(row_cells,_col_ai)
            nh = _cell_val(row_cells, _col_nh) if _col_nh else ""
            if not ic and not in_ and not sd and not dv and not ev: continue
            raw_rows.append({"row_idx":row_idx,"source_item_code":ic,"source_item_name":in_,
                              "source_desc":sd,"existing_item_name_vi":dv,
                              "existing_desc_vi":ev,"existing_ai":ai,"source_nhan_hieu":nh})
        wb_ro.close()
        print(f"Tổng dòng đọc được: {len(raw_rows)}")

        # Xác định hạng mục lá (leaf):
        #   - Số La Mã thuần (I, II, III...) → luôn là phân mục cha
        #   - Code khác → lá nếu không có code nào bắt đầu bằng code + "."
        _all_codes_xl = [r["source_item_code"] for r in raw_rows if r["source_item_code"]]
        _leaf_codes_xl = {c for c in _all_codes_xl
                          if not is_roman_numeral_code(c)
                          and not any(o.startswith(c + ".") for o in _all_codes_xl if o != c)}
        _non_leaf_count = sum(1 for r in raw_rows if r["source_item_code"] and r["source_item_code"] not in _leaf_codes_xl)
        print(f"  Hạng mục lá (cần sinh mô tả): {len(_leaf_codes_xl)}, hạng mục cha (bỏ qua mô tả): {_non_leaf_count}")

        def _get_ws_write():
            nonlocal _wb_write, _ws_write
            if _wb_write is None:
                print("  [LAZY] Load workbook để ghi...")
                _wb_write = load_workbook(readable_input, keep_vba=keep_vba)
                _sheet    = SHEET_NAME if (SHEET_NAME and SHEET_NAME in _wb_write.sheetnames) else _wb_write.sheetnames[0]
                _ws_write = _wb_write[_sheet]
                setup_columns(_ws_write, col_map)
            return _ws_write

    pending_rows: List[dict] = []

    for r in raw_rows:
        source_item_name    = r["source_item_name"]
        existing_ai         = r["existing_ai"]
        source_desc         = r["source_desc"]
        source_desc_cleaned = remove_duplicate_tail_item(source_desc, source_item_name)

        # Cột B (item name) thường ngắn → dùng is_foreign_short (không có dấu VN = ngoại ngữ)
        # Cột C (desc) dài → dùng still_has_foreign (đếm tỉ lệ từ)
        _b_is_foreign = is_foreign_short(source_item_name) and not is_technical_code(source_item_name)
        _c_is_foreign = still_has_foreign(source_desc_cleaned) and not is_technical_code(source_desc_cleaned)

        _need_translate_D = _b_is_foreign
        _need_clear_D     = (not _b_is_foreign) and bool(normalize_inline(r["existing_item_name_vi"]))
        _need_translate_E = _c_is_foreign
        _need_clear_E     = (not _c_is_foreign) and bool(normalize_spaces(r["existing_desc_vi"]))

        # Chỉ sinh AI desc cho hạng mục lá; hạng mục cha (có con) bỏ qua
        _ic = r["source_item_code"]
        if not is_json_input:
            _is_leaf = (not _ic) or (_ic in _leaf_codes_xl)
        else:
            _is_leaf = True  # JSON đã lọc non-leaf từ trước
        need_ai      = _is_leaf and ((not normalize_spaces(existing_ai)) or FORCE_REPROCESS_AI)
        need_clear_ai = (not _is_leaf) and bool(normalize_spaces(existing_ai))

        item_name_vi_seed       = "" if _b_is_foreign else (r["existing_item_name_vi"] if looks_like_vietnamese(r["existing_item_name_vi"]) else "")
        translated_desc_vi_seed = "" if _c_is_foreign else (r["existing_desc_vi"] if looks_like_vietnamese(r["existing_desc_vi"]) else "")

        # Luôn đi qua pipeline đầy đủ: dịch → lưu cache → similarity → AI
        # TRANSLATION_CACHE và AI_DESC_CACHE sẽ tự skip nếu đã có kết quả
        pending_rows.append({
            "row_idx": r["row_idx"], "source_item_code": r["source_item_code"],
            "source_item_name": source_item_name, "source_desc": source_desc,
            "source_desc_cleaned": source_desc_cleaned,
            "existing_item_name_vi": r["existing_item_name_vi"],
            "existing_desc_vi": r["existing_desc_vi"], "existing_ai": existing_ai,
            "item_name_vi_seed": item_name_vi_seed,
            "translated_desc_vi_seed": translated_desc_vi_seed,
            "need_item_name_vi": _need_translate_D,
            "need_desc_vi": _need_translate_E,
            "need_clear_item_name_vi": _need_clear_D,
            "need_clear_desc_vi": _need_clear_E,
            "need_ai": need_ai,
            "need_clear_ai": need_clear_ai,
            "source_nhan_hieu": r.get("source_nhan_hieu", ""),
        })

    print(f"Cần xử lý: {len(pending_rows)}")
    if not pending_rows:
        save_persistent_state(force=True); clear_checkpoint()
        print("Không có dòng nào cần xử lý. Xong."); return

    _prepop_count = 0
    for row in pending_rows:
        if row.get("need_item_name_vi") or row.get("need_desc_vi"): continue
        if "final_item_name_vi" in row: continue
        row["final_item_name_vi"] = row["existing_item_name_vi"] or \
            fallback_translate_item_name(row["source_item_name"], row["source_desc_cleaned"])
        row["final_desc_vi"]      = row["existing_desc_vi"] or \
            (fallback_translate_desc(row["source_desc_cleaned"], row["source_item_name"])
             if not looks_like_vietnamese(row["source_desc_cleaned"]) else
             normalize_vi_text(row["source_desc_cleaned"]))
        _prepop_count += 1
    if _prepop_count:
        print(f"  Pre-populate {_prepop_count} rows không cần dịch")

    all_batches  = list(chunk_list(pending_rows, BATCH_SIZE))
    t0           = time.time()
    all_processed= asyncio.run(process_all_batches_async(all_batches))
    print(f"  ⏱ API pipeline: {time.time()-t0:.1f}s cho {len(all_processed)} rows")

    batches_processed = list(chunk_list(all_processed, BATCH_SIZE))
    done = 0
    pbar_write = _tqdm(total=len(all_processed), desc="Ghi JSON" if is_json_input else "Ghi Excel", unit="rows")

    for idx, batch in enumerate(batches_processed, start=1):
        for row in batch:
            row_idx            = row["row_idx"]
            final_item_name_vi = row.get("final_item_name_vi","")
            final_desc_vi      = row.get("final_desc_vi","")
            final_ai           = row.get("final_ai","")
            _final_ai_orig     = final_ai
            _nh_val            = row.get("source_nhan_hieu", "").strip()
            if _nh_val and _nh_val.lower() not in final_ai.lower():
                _base   = final_ai.rstrip()
                if _base and _base[-1] not in ".!?":
                    _base += "."
                final_ai = (_base + " " if _base else "") + f"Nhãn hiệu: {_nh_val}."
            elif not _nh_val:
                # JSON input: brand embedded in source_desc
                final_ai = _append_brand_to_ai_desc(final_ai, row.get("source_desc_cleaned",""), "")
            _brand_appended    = final_ai != _final_ai_orig

            if is_json_input:
                item = _json_idx_map.get(row_idx)
                if item is not None:
                    if row.get("need_item_name_vi"):
                        item["translated_item_name"] = final_item_name_vi
                    elif row.get("need_clear_item_name_vi"):
                        item["translated_item_name"] = ""
                    if row.get("need_desc_vi"):
                        item["translated_desc"] = final_desc_vi
                    elif row.get("need_clear_desc_vi"):
                        item["translated_desc"] = ""
                    if row.get("need_ai") or _brand_appended:
                        item["ai_desc"] = final_ai
            else:
                # Ghi D (B→D)
                if row.get("need_item_name_vi"):
                    print(f"  [WRITE D] row {row_idx}: '{final_item_name_vi[:50]}'")
                    write_cell(_get_ws_write(), row_idx, col_map["translated_item_name"], final_item_name_vi)
                elif row.get("need_clear_item_name_vi") and col_map["translated_item_name"]:
                    write_cell(_get_ws_write(), row_idx, col_map["translated_item_name"], "")

                # Ghi E (C→E)
                if row.get("need_desc_vi"):
                    print(f"  [WRITE E] row {row_idx}: '{final_desc_vi[:50]}'" if final_desc_vi else f"  [EMPTY E] row {row_idx}: final_desc_vi RỖNG!")
                    write_cell(_get_ws_write(), row_idx, col_map["translated_desc"], final_desc_vi)
                elif row.get("need_clear_desc_vi") and col_map["translated_desc"]:
                    write_cell(_get_ws_write(), row_idx, col_map["translated_desc"], "")

                # Ghi F
                if row.get("need_ai") or _brand_appended:
                    write_cell(_get_ws_write(), row_idx, col_map["ai_desc"], final_ai)
                elif row.get("need_clear_ai") and col_map["ai_desc"]:
                    write_cell(_get_ws_write(), row_idx, col_map["ai_desc"], "")

                if _wb_write:
                    _get_ws_write().row_dimensions[row_idx].height = estimate_row_height(
                        final_item_name_vi, final_desc_vi, final_ai)

            _name_quality_ok = (not row.get("need_item_name_vi")) or looks_like_vietnamese(final_item_name_vi)
            _desc_quality_ok = (not row.get("need_desc_vi")) or looks_like_vietnamese(final_desc_vi)
            if _name_quality_ok and _desc_quality_ok and (normalize_spaces(final_ai) or final_item_name_vi):
                _out_key = get_output_cache_key(row["source_item_name"], row["source_desc_cleaned"])
                OUTPUT_CACHE[_out_key] = {"item_name_vi":final_item_name_vi,"desc_vi":final_desc_vi,"ai_desc":final_ai}
                _INST_OUTPUT_CACHE_KEYS.add(_out_key)
                global _OUTPUT_CACHE_DIRTY; _OUTPUT_CACHE_DIRTY = True

            done += 1
            pbar_write.update(1)

        if is_json_input:
            if idx % SAVE_EVERY_N_BATCHES == 0 or idx == len(batches_processed):
                write_json_file_atomic(Path(INPUT_FILE), {"data": _json_data})
        elif _wb_write and (idx % SAVE_EVERY_N_BATCHES == 0 or idx == len(batches_processed)):
            _safe_wb_save(_wb_write, OUTPUT_FILE, f"batch {idx}/{len(batches_processed)}")

        fresh_rows = [r for r in batch if not r.get("_from_output_cache")]
        learned    = auto_learn_from_batch(fresh_rows)
        save_persistent_state()
        if batch: save_checkpoint(INPUT_FILE, batch[-1]["row_idx"])
        maybe_refresh_knowledge_index_after_learning(learned)

    pbar_write.close()
    if is_json_input:
        write_json_file_atomic(Path(INPUT_FILE), {"data": _json_data})
        print(f"Xong. Output JSON: {INPUT_FILE}")
    elif _wb_write:
        _safe_wb_save(_wb_write, OUTPUT_FILE, "final")
    else:
        print("Không cần ghi Excel — tất cả đã đúng.")

    # Giải phóng GPU memory khi xong
    if HAS_TORCH and DEVICE == "cuda":
        torch.cuda.empty_cache()
        print(f"  [GPU] VRAM released. Peak usage: {torch.cuda.max_memory_allocated()/1e9:.2f}GB")

    save_persistent_state(force=True); clear_checkpoint()
    print(f"\n✓ Xong. Output: {OUTPUT_FILE}")
    print(f"  Device:            {DEVICE}")
    print(f"  Done rows:         {done}")
    print(f"  Translation cache: {len(TRANSLATION_CACHE)}")
    print(f"  AI desc cache:     {len(AI_DESC_CACHE)}")
    print(f"  Similarity cache:  {len(SIM_AUDIT_CACHE)}")
    print(f"  Output cache:      {len(OUTPUT_CACHE)}")
    print(f"  Learned knowledge: {len(LEARNED_ITEMS_RAM)}")


if __name__ == "__main__":
    main()